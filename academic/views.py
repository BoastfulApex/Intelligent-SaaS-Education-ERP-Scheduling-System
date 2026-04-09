import re
import pandas as pd
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from django.db import transaction
from .models import (Major, Subject, Curriculum, CurriculumBlock,
                     CurriculumSubject, Group, Shift, Para, GroupAssignment)
from .serializers import (MajorSerializer, SubjectSerializer,
                           CurriculumSerializer, CurriculumBlockSerializer,
                           CurriculumSubjectSerializer, GroupSerializer,
                           ShiftSerializer, ParaSerializer,
                           GroupAssignmentSerializer)
from permissions import IsEduAdmin, IsOrgAdmin, IsOrgAdminOrReadOnly, IsEduAdminOrReadOnly


# ──────────────────────────────────────────────────────────────────────────────
#  O'QUV REJA EXCEL PARSER
# ──────────────────────────────────────────────────────────────────────────────

ROMAN = {'I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII', 'IX', 'X'}

def _is_block_row(tr_val) -> bool:
    """I., II., III., IV. — blok qatori."""
    if pd.isna(tr_val):
        return False
    s = str(tr_val).strip().rstrip('.')
    return s in ROMAN

def _is_subject_row(tr_val) -> bool:
    """1.1., 2.3. — fan qatori."""
    if pd.isna(tr_val):
        return False
    s = str(tr_val).strip()
    return bool(re.match(r'^\d+\.\d+\.?$', s))

def _is_total_row(tr_val) -> bool:
    """Jami: — yakuniy qator."""
    if pd.isna(tr_val):
        return False
    return 'jami' in str(tr_val).strip().lower()

def _safe_int(val) -> int:
    """NaN yoki noto'g'ri qiymatni 0 ga aylantiradi."""
    try:
        if pd.isna(val):
            return 0
        return int(float(val))
    except (TypeError, ValueError):
        return 0

def _extract_meta(df, row_hint: int, prefix: str) -> str:
    """
    'Tinglovchilar kontingenti: Sport...' kabi qatordan qiymatni ajratadi.
    Avval row_hint qatorida, topilmasa barcha qatorlarda qidiradi.
    """
    def _check_row(r_idx):
        for col in range(df.shape[1]):
            val = df.iloc[r_idx, col]
            if pd.notna(val) and prefix.lower() in str(val).lower():
                text = str(val).strip()
                # Qiymat ikkinchi ustunda bo'lishi mumkin
                next_col = col + 1
                if ':' in text:
                    right = text.split(':', 1)[1].strip()
                    if right:
                        return right
                    # Bo'sh bo'lsa keyingi ustundan ol
                    if next_col < df.shape[1]:
                        nv = df.iloc[r_idx, next_col]
                        if pd.notna(nv):
                            return str(nv).strip()
                else:
                    # Keyingi ustun qiymatini qaytarish
                    if next_col < df.shape[1]:
                        nv = df.iloc[r_idx, next_col]
                        if pd.notna(nv):
                            return str(nv).strip()
                    return text
        return ''

    # Avval ko'rsatilgan qatorda qidirish (±3 qator oraliq)
    for offset in range(-3, 4):
        r = row_hint + offset
        if 0 <= r < df.shape[0]:
            result = _check_row(r)
            if result:
                return result

    # Topilmasa — barcha qatorlarni qidirish
    for r in range(df.shape[0]):
        result = _check_row(r)
        if result:
            return result
    return ''


def _find_data_start(df) -> int:
    """Birinchi blok qatorini topadi (I. bor qator).
    0-ustun bo'sh bo'lsa 1-ustunni ham tekshiradi."""
    for i in range(df.shape[0]):
        if _is_block_row(df.iloc[i, 0]):
            return i
        # Ba'zi fayllarda 0-ustun bo'sh, blok 1-ustunda
        if df.shape[1] > 1 and _is_block_row(df.iloc[i, 1]):
            return i
    return -1

def _parse_study_form(text: str) -> str:
    """O'qish shakli textdan StudyForm choicega."""
    t = text.lower()
    if 'sirtqi' in t or 'ajralmagan' in t:
        return Curriculum.StudyForm.PARTTIME
    if 'masofaviy' in t or 'online' in t or 'onlayn' in t:
        return Curriculum.StudyForm.DISTANCE
    return Curriculum.StudyForm.FULLTIME

def _parse_duration(text: str) -> tuple[int, int]:
    """'4 hafta (144 soat)' → (4, 144)."""
    weeks = hours = 0
    m = re.search(r'(\d+)\s*hafta', text)
    if m:
        weeks = int(m.group(1))
    m = re.search(r'(\d+)\s*soat', text)
    if m:
        hours = int(m.group(1))
    return weeks, hours

def _generate_subject_code(name: str) -> str:
    """
    Fan nomining har bir so'zidan birinchi harf olib kod yasaydi.
    "Jismoniy tayyorgarlik"  → "JT"
    "Sport mashg'uloti"      → "SM"
    "Umumiy o'rta ta'lim"    → "UOT"
    Kichik xizmat so'zlari (va, va, bilan, ...) o'tkazib yuboriladi.
    """
    STOP_WORDS = {"va", "bilan", "uchun", "bu", "bir", "ham", "yoki", "bo'yicha"}
    words = re.split(r"[\s\-–—]+", name.strip())
    initials = []
    for w in words:
        clean = re.sub(r"[^a-zA-ZА-Яа-яёЁA-Za-zʻʼ']", '', w)
        if clean and clean.lower() not in STOP_WORDS:
            initials.append(clean[0].upper())
    return ''.join(initials) if initials else name[:4].upper()


def _get_or_create_subject(code: str, name: str, organization) -> tuple:
    """
    Fan topish yoki yaratish. Qaytaradi: (Subject, created: bool)

    Qidiruv tartibi:
      1. Nom bo'yicha aniq moslik (case-insensitive) — mavjud fanni qayta ishlatish
      2. Topilmasa — yangi Subject yaratish
         Kod = fan nomining so'zlari bosh harflaridan (masalan "JT", "SM")
         Noyobligini ta'minlash uchun raqam qo'shiladi: "JT", "JT-2", ...
    """
    clean_name = name.strip()

    # 1. Nom bo'yicha qidirish
    by_name = Subject.objects.filter(
        organization=organization,
        name__iexact=clean_name,
    ).first()
    if by_name:
        return by_name, False

    # 2. Yangi fan yaratish — kod nomning bosh harflaridan
    base_code  = _generate_subject_code(clean_name)
    final_code = base_code
    counter    = 2
    while Subject.objects.filter(organization=organization, code=final_code).exists():
        final_code = f"{base_code}-{counter}"
        counter += 1

    subj = Subject.objects.create(
        organization=organization,
        code=final_code,
        name=clean_name,
    )
    return subj, True


def parse_curriculum_excel(file, major: Major, organization,
                           curriculum_name: str = '') -> dict:
    """
    O'quv reja Excel faylini o'qib, Curriculum + Block + Subject larni
    yaratadi yoki yangilaydi.

    Qaytaradi:
        { 'curriculum': Curriculum, 'blocks': int, 'subjects': int,
          'warnings': list[str] }
    """
    warnings = []
    df = pd.read_excel(file, sheet_name=0, header=None)

    # ── METADATA ──────────────────────────────────────────────────────────────
    # User kiritgan nom ustuvorlik — agar bo'sh bo'lsa Exceldan olamiz
    if curriculum_name:
        # Foydalanuvchi kiritgan nom — Excelga ustuvorlik
        course_name = curriculum_name
    else:
        # Exceldan aniq "Kurs nomi:" ni qidirish
        course_name = (
            _extract_meta(df, 13, 'kurs nomi') or  # "Kurs nomi: ..." aniq prefix
            _extract_meta(df, 13, 'kurs')           # keng fallback
        )
        # Hali ham topilmasa — yo'nalish nomidan yasaymiz
        if not course_name:
            course_name = f"{major.name} O'quv reja"

    contingent   = _extract_meta(df, 14, 'kontingenti')
    duration_txt = _extract_meta(df, 15, 'muddati')
    form_txt     = _extract_meta(df, 16, 'shakli')

    duration_weeks, total_hours = _parse_duration(duration_txt)
    study_form = _parse_study_form(form_txt)

    # ── DATA QATORLAR BOSHLANISH NUQTASI ──────────────────────────────────────
    data_start = _find_data_start(df)
    if data_start < 0:
        return {
            'curriculum': None,
            'blocks': 0, 'subjects': 0,
            'warnings': ["Excel faylda blok qatorlari (I., II., ...) topilmadi."],
        }

    # ── CURRICULUM YARATISH (mavjud bo'lsa arxivlash) ─────────────────────────
    with transaction.atomic():
        # Avvalgi faol reja arxivlash
        Curriculum.objects.filter(
            major=major,
            status=Curriculum.Status.ACTIVE,
        ).update(status=Curriculum.Status.ARCHIVED)

        curriculum = Curriculum.objects.create(
            major=major,
            name=course_name,
            contingent=contingent,
            study_form=study_form,
            duration_weeks=duration_weeks or 4,
            total_hours=total_hours or 144,
            status=Curriculum.Status.ACTIVE,
        )

        current_block   = None
        block_order     = 0
        subject_order   = 0
        blocks_count    = 0
        subjects_count  = 0
        subjects_new    = 0   # yangi yaratilgan
        subjects_linked = 0   # mavjud topilgan va bog'langan

        # Col offset: ba'zi fayllarda 0-ustun bo'sh, hamma narsa 1 ga siljigan
        col_offset = 0
        if df.shape[1] > 1 and _is_block_row(df.iloc[data_start, 1]):
            col_offset = 1

        for row_idx in range(data_start, df.shape[0]):
            row      = df.iloc[row_idx]
            ncols    = len(row)
            tr_val   = row.iloc[col_offset] if ncols > col_offset else None
            name_col = col_offset + 1
            name_val = str(row.iloc[name_col]).strip() if ncols > name_col and pd.notna(row.iloc[name_col]) else ''

            # ── JAMI qatori — to'xtatish ──────────────────────────────────────
            if _is_total_row(tr_val):
                break
            # Bo'sh qator — o'tkazib yuborish
            if tr_val is None or (pd.isna(tr_val) and not name_val):
                continue

            # ── BLOK qatori ───────────────────────────────────────────────────
            if _is_block_row(tr_val):
                block_order += 1
                subject_order = 0
                # Blok nomi: name_val dan olinadi (bo'lsa)
                block_name = name_val if name_val else f'Blok {block_order}'
                current_block = CurriculumBlock.objects.create(
                    curriculum=curriculum,
                    name=block_name,
                    order=block_order,
                )
                blocks_count += 1
                continue

            # ── FAN qatori ────────────────────────────────────────────────────
            if _is_subject_row(tr_val) and current_block and name_val:
                subject_order += 1
                code = str(tr_val).strip().rstrip('.')

                # Col indekslari (col_offset hisobga olingan):
                # 0=T/r, 1=Fan nomi, 2=Jami, 3=Nazariy(ma'ruza), 4=Amaliy,
                # 5=Ko'chma, 6=Mustaqil, 7=I-hafta, 8=II-h, 9=III-h, 10=IV-h
                # Asosiy format (col_offset=0): 6=Leksiya, 7=Amaliy, 8=Ko'chma
                def _gc(ci):
                    """col_offset qo'shib xavfsiz olish."""
                    idx = ci + col_offset
                    return row.iloc[idx] if ncols > idx else None

                lecture_h     = _safe_int(_gc(6))
                practice_h    = _safe_int(_gc(7))
                field_h       = _safe_int(_gc(8))
                independent_h = _safe_int(_gc(9))
                week1_h       = _safe_int(_gc(10))
                week2_h       = _safe_int(_gc(11))
                week3_h       = _safe_int(_gc(12))
                week4_h       = _safe_int(_gc(13))

                subject, created = _get_or_create_subject(code, name_val, organization)

                CurriculumSubject.objects.create(
                    block=current_block,
                    subject=subject,
                    order=subject_order,
                    lecture_hours=lecture_h,
                    practice_hours=practice_h,
                    field_hours=field_h,
                    independent_hours=independent_h,
                    week1_hours=week1_h,
                    week2_hours=week2_h,
                    week3_hours=week3_h,
                    week4_hours=week4_h,
                )
                subjects_count += 1
                if created:
                    subjects_new += 1
                else:
                    subjects_linked += 1

    return {
        'curriculum':     curriculum,
        'blocks':         blocks_count,
        'subjects':       subjects_count,
        'subjects_new':   subjects_new,
        'subjects_linked': subjects_linked,
        'warnings':       warnings,
    }


class MajorViewSet(viewsets.ModelViewSet):
    serializer_class = MajorSerializer
    permission_classes = [IsEduAdmin]

    def get_queryset(self):
        return Major.objects.filter(
            organization=self.request.user.organization
        )

    def perform_create(self, serializer):
        serializer.save(organization=self.request.user.organization)


class SubjectViewSet(viewsets.ModelViewSet):
    serializer_class = SubjectSerializer
    permission_classes = [IsEduAdmin]

    def get_queryset(self):
        return Subject.objects.filter(
            organization=self.request.user.organization
        )
                
    @action(detail=False, methods=['post'], url_path='bulk-create')
    def bulk_create(self, request):
        """Bir vaqtda ko'p fan yaratish"""
        subjects = request.data.get('subjects', [])
        if not subjects:
            return Response(
                {'error': 'subjects maydoni bo\'sh!'},
                status=status.HTTP_400_BAD_REQUEST
            )

        created = []
        errors  = []

        for item in subjects:
            try:
                subject, _ = Subject.objects.get_or_create(
                    organization=request.user.organization,
                    code=item['code'],
                    defaults={
                        'name': item['name'],
                        'department_id': item.get('department')  # None bo'lsa ham ishlaydi
                    }
                )
                created.append(subject)
            except Exception as e:
                errors.append({'code': item.get('code'), 'error': str(e)})

        return Response({
            'created': SubjectSerializer(created, many=True).data,
            'errors': errors
        }, status=status.HTTP_201_CREATED)
    
    def perform_create(self, serializer):
        serializer.save(organization=self.request.user.organization)


class CurriculumViewSet(viewsets.ModelViewSet):
    serializer_class   = CurriculumSerializer
    permission_classes = [IsEduAdmin]
    parser_classes     = [MultiPartParser, FormParser]

    def get_queryset(self):
        return Curriculum.objects.filter(
            major__organization=self.request.user.organization
        ).prefetch_related('blocks__subjects')

    @action(detail=True, methods=['post'])
    def archive(self, request, pk=None):
        curriculum = self.get_object()
        if curriculum.status == Curriculum.Status.ARCHIVED:
            return Response(
                {'error': 'Bu o\'quv reja allaqachon arxivlangan!'},
                status=status.HTTP_400_BAD_REQUEST
            )
        curriculum.archive()
        return Response({'message': 'O\'quv reja arxivlandi'})

    @action(detail=False, methods=['post'], url_path='upload',
            parser_classes=[MultiPartParser, FormParser])
    def upload(self, request):
        """
        POST /api/v1/curriculums/upload/
        Body (form-data):
          - file:      Excel fayl (.xlsx)  — O'quv reja
          - major_id:  Yo'nalish ID si
          - name:      O'quv reja nomi (ixtiyoriy, Excel dan olinadi)

        Logika:
          - Avvalgi faol reja arxivlanadi
          - Yangi Curriculum, CurriculumBlock, CurriculumSubject lar yaratiladi
          - Subject lar code bo'yicha topiladi yoki yangi yaratiladi
          - CurriculumBlock.department = null (edu admin keyinchalik biriktiradi)
        """
        file = request.FILES.get('file')
        if not file:
            return Response(
                {'error': '"file" maydoni talab qilinadi.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'Faqat .xlsx yoki .xls fayl qabul qilinadi.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        major_id = request.data.get('major_id')
        if not major_id:
            return Response(
                {'error': '"major_id" maydoni talab qilinadi.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        major = Major.objects.filter(
            id=major_id,
            organization=request.user.organization,
        ).first()
        if not major:
            return Response(
                {'error': 'Yo\'nalish topilmadi.'},
                status=status.HTTP_404_NOT_FOUND
            )

        curriculum_name = request.data.get('name', '')

        try:
            result = parse_curriculum_excel(
                file=file,
                major=major,
                organization=request.user.organization,
                curriculum_name=curriculum_name,
            )
        except Exception as e:
            return Response(
                {'error': f'Excel faylni o\'qishda xato: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if result['curriculum'] is None:
            return Response(
                {'error': result['warnings'][0] if result['warnings'] else 'Noma\'lum xato.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        return Response({
            'success':    True,
            'message':    (
                f"O'quv reja yuklandi: {result['blocks']} blok, "
                f"{result['subjects']} fan "
                f"({result['subjects_new']} yangi, "
                f"{result['subjects_linked']} mavjud bog'landi)."
            ),
            'curriculum': CurriculumSerializer(result['curriculum']).data,
            'stats': {
                'blocks':          result['blocks'],
                'subjects':        result['subjects'],
                'subjects_new':    result['subjects_new'],
                'subjects_linked': result['subjects_linked'],
            },
            'warnings': result['warnings'],
        }, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'], url_path='template',
            permission_classes=[IsAuthenticated])
    def template(self, request):
        """
        GET /api/v1/curriculums/template/
        O'quv reja Excel shablonini yuklab beradi.
        """
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "O'quv reja"

        blue_fill   = PatternFill('solid', fgColor='1F4E79')
        gray_fill   = PatternFill('solid', fgColor='D9D9D9')
        block_fill  = PatternFill('solid', fgColor='BDD7EE')
        white_font  = Font(bold=True, color='FFFFFF', size=10)
        black_bold  = Font(bold=True, size=10)
        center      = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left        = Alignment(horizontal='left',   vertical='center', wrap_text=True)
        thin        = Side(style='thin', color='AAAAAA')
        border      = Border(left=thin, right=thin, top=thin, bottom=thin)

        def cell(r, c, val='', fill=None, font=None, align=None, bold=False):
            cl = ws.cell(row=r, column=c, value=val)
            if fill:  cl.fill      = fill
            if font:  cl.font      = font
            elif bold: cl.font     = Font(bold=True, size=10)
            if align: cl.alignment = align
            cl.border = border
            return cl

        # ── METADATA qatorlar (1-16) ──────────────────────────────────────────
        meta_rows = [
            (1,  "SPORT TA'LIM MUASSASASI"),
            (2,  ""),
            (10, "Kurs nomi:"),
            (11, "Yo'nalish:"),
            (12, "Tinglovchilar kontingenti: Sport...",),
            (13, "Kurs nomi: Jismoniy tarbiya va sport mutaxassisligiga kirish"),
            (14, "Tinglovchilar kontingenti: Sport mutaxassisligi"),
            (15, "O'qish muddati: 4 hafta (144 soat)"),
            (16, "O'qish shakli: Kunduzgi"),
        ]
        for r, txt in meta_rows:
            ws.cell(row=r, column=1, value=txt).font = Font(bold=True, size=10)

        ws.merge_cells('A1:N1')
        ws.cell(row=1, column=1).alignment = center

        # ── JADVAL SARLAVHASI (17-qator) ─────────────────────────────────────
        headers = [
            'T/r', 'Fan nomi / Modul', 'Jami soat',
            "Ma'ruza (Nazariy)", 'Amaliy', "Ko'chma", 'Mustaqil',
            'I-hafta', 'II-hafta', 'III-hafta', 'IV-hafta',
            'Izoh',
        ]
        for c, h in enumerate(headers, start=1):
            cell(17, c, h, fill=blue_fill, font=white_font, align=center)

        ws.row_dimensions[17].height = 45

        # ── MISOL MA'LUMOTLAR ─────────────────────────────────────────────────
        example_data = [
            # (T/r, Fan nomi, Jami, Leksiya, Amaliy, Ko'chma, Mustaqil, w1, w2, w3, w4, Izoh)
            ('I.',   'JISMONIY TAYYORGARLIK BLOKI',  '',  '', '', '', '', '', '', '', '', ''),
            ('1.1.', 'Jismoniy tayyorgarlik',          40, 16,  8,  8,  8, 10, 10, 10, 10, ''),
            ('1.2.', 'Sport mashg\'uloti',              40,  8, 16,  8,  8, 10, 10, 10, 10, ''),
            ('1.3.', 'Maxsus jismoniy tayyorgarlik',   24,  8,  8,  4,  4,  6,  6,  6,  6, ''),
            ('II.',  'NAZARIY BILIMLAR BLOKI',          '',  '', '', '', '', '', '', '', '', ''),
            ('2.1.', 'Nazariya va metodika',            16, 12,  4,  0,  0,  4,  4,  4,  4, ''),
            ('2.2.', 'Taktika',                         16,  8,  8,  0,  0,  4,  4,  4,  4, ''),
            ('III.', 'AMALIY KO\'NIKMALAR BLOKI',       '',  '', '', '', '', '', '', '', '', ''),
            ('3.1.', 'Pedagogik amaliyot',               8,  0,  8,  0,  0,  2,  2,  2,  2, ''),
            ('Jami:', '',                               144, 52, 52, 20, 20, 36, 36, 36, 36, ''),
        ]

        for r_offset, row_data in enumerate(example_data):
            r = 18 + r_offset
            tr = row_data[0]
            is_block = _is_block_row(tr)
            is_total = _is_total_row(tr)
            bg = block_fill if is_block else (gray_fill if is_total else None)
            fn = black_bold if (is_block or is_total) else Font(size=10)
            for c, val in enumerate(row_data, start=1):
                cl = cell(r, c, val if val != '' else None, fill=bg, font=fn,
                          align=center if c != 2 else left)

        # ── USTUN KENGLIKLARI ─────────────────────────────────────────────────
        col_widths = [6, 35, 8, 14, 8, 8, 9, 8, 8, 8, 8, 15]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

        ws.freeze_panes = 'A18'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = 'attachment; filename="oquv_reja_shablon.xlsx"'
        return resp


class CurriculumBlockViewSet(viewsets.ModelViewSet):
    serializer_class = CurriculumBlockSerializer
    permission_classes = [IsEduAdmin]

    def get_queryset(self):
        qs = CurriculumBlock.objects.filter(
            curriculum__major__organization=self.request.user.organization
        ).select_related('curriculum').prefetch_related('subjects__department', 'subjects__subject')

        curriculum_id = self.request.query_params.get('curriculum_id')
        if curriculum_id:
            qs = qs.filter(curriculum_id=curriculum_id)
        return qs


class CurriculumSubjectViewSet(viewsets.ModelViewSet):
    serializer_class = CurriculumSubjectSerializer
    permission_classes = [IsEduAdmin]

    def get_queryset(self):
        qs = CurriculumSubject.objects.filter(
            block__curriculum__major__organization=self.request.user.organization
        ).select_related('subject', 'block', 'department')

        curriculum_id = self.request.query_params.get('curriculum_id')
        if curriculum_id:
            qs = qs.filter(block__curriculum_id=curriculum_id)

        block_id = self.request.query_params.get('block_id')
        if block_id:
            qs = qs.filter(block_id=block_id)

        unassigned = self.request.query_params.get('unassigned')
        if unassigned in ('1', 'true'):
            qs = qs.filter(department__isnull=True)

        return qs

    @action(detail=False, methods=['post'], url_path='assign-departments')
    def assign_departments(self, request):
        """
        Fanlarga kafedra biriktirish.

        Body:
        {
          "assignments": [
            {"subject_id": 12, "department_id": 3},
            {"subject_id": 13, "department_id": 5},
            {"subject_id": 14, "department_id": null}   ← olib tashlash
          ]
        }
        """
        from organizations.models import Department

        assignments = request.data.get('assignments', [])
        if not assignments:
            return Response({'error': '"assignments" bo\'sh'}, status=400)

        org     = request.user.organization
        updated = []
        errors  = []

        with transaction.atomic():
            for item in assignments:
                subj_id = item.get('subject_id')
                dept_id = item.get('department_id')

                cs = CurriculumSubject.objects.filter(
                    id=subj_id,
                    block__curriculum__major__organization=org,
                ).first()
                if not cs:
                    errors.append({'subject_id': subj_id, 'error': 'Fan topilmadi'})
                    continue

                if dept_id:
                    dept = Department.objects.filter(id=dept_id, organization=org).first()
                    if not dept:
                        errors.append({'subject_id': subj_id, 'error': 'Kafedra topilmadi'})
                        continue
                    cs.department = dept
                else:
                    cs.department = None

                cs.save(update_fields=['department'])
                updated.append(CurriculumSubjectSerializer(cs).data)

        return Response({'updated': updated, 'errors': errors})


class GroupViewSet(viewsets.ModelViewSet):
    serializer_class = GroupSerializer
    permission_classes = [IsEduAdmin]

    def get_queryset(self):
        return Group.objects.filter(
            organization=self.request.user.organization,
            is_active=True
        )

    def perform_create(self, serializer):
        serializer.save(organization=self.request.user.organization)


class ShiftViewSet(viewsets.ModelViewSet):
    serializer_class = ShiftSerializer
    permission_classes = [IsOrgAdmin]

    def get_queryset(self):
        return Shift.objects.filter(
            organization=self.request.user.organization,
            is_active=True
        ).prefetch_related('paras')

    def perform_create(self, serializer):
        serializer.save(organization=self.request.user.organization)


class ParaViewSet(viewsets.ModelViewSet):
    serializer_class = ParaSerializer
    permission_classes = [IsOrgAdmin]

    def get_queryset(self):
        return Para.objects.filter(
            shift__organization=self.request.user.organization,
            is_active=True
        ).select_related('shift')


class GroupAssignmentViewSet(viewsets.ModelViewSet):
    serializer_class = GroupAssignmentSerializer
    permission_classes = [IsEduAdmin]

    def get_queryset(self):
        return GroupAssignment.objects.filter(
            group__organization=self.request.user.organization
        ).select_related('group', 'shift', 'building')