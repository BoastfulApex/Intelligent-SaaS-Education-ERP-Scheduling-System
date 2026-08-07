from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
import threading
from django.db import connection as db_connection
from django.utils import timezone

# Jadval tuzish odatda 4-6 daqiqa. Shundan ancha ko'p `running` holatda turgan
# jarayon - o'lib ketgan oqim (server qayta ishga tushgan va h.k.). Uni xato
# deb belgilamasak, qayta generatsiya abadiy bloklanib qolardi (409).
STALE_GEN_SECONDS = 1800
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.pagination import PageNumberPagination
from django.db import transaction
from django.db.models import Q, Count
from django.http import HttpResponse
import calendar
import datetime
import io
import logging
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd
from permissions import IsDeptManager, IsEduAdmin, IsOrgAdmin, IsDeptManagerOrReadOnly

logger = logging.getLogger(__name__)

from .models import (Teacher, TeacherBusyTime, TeacherSubjectAssignment,
                     TeacherMonthlyLoad, Schedule, ScheduleEntry,
                     Substitution, AuditLog,
                     LoadSheet, TeacherLoad, LoadDistribution,
                     GroupSubject)
from .serializers import (TeacherSerializer, TeacherBusyTimeSerializer,
                           TeacherSubjectAssignmentSerializer,
                           TeacherMonthlyLoadSerializer, ScheduleSerializer,
                           ScheduleEntrySerializer, SubstitutionSerializer,
                           AuditLogSerializer,
                           LoadSheetSerializer)
from academic.models import Para, GroupAssignment, Group, CurriculumSubject, Curriculum, DeliveryMode
from organizations.models import Room, Department


# ─── TAQSIMOT PARSER YORDAMCHI FUNKSIYALARI ───────────────────────────────────

MONTH_MAP = {
    'yanvar': 1, 'fevral': 2, 'mart': 3, 'aprel': 4,
    'may': 5, 'iyun': 6, 'iyul': 7, 'avgust': 8,
    'sentabr': 9, 'sentyabr': 9,   # ikkala imlo ham uchraydi
    'oktabr': 10, 'noyabr': 11, 'dekabr': 12,
}

STAVKA_MAP = {
    '1':    TeacherLoad.Stavka.FULL,
    '1.0':  TeacherLoad.Stavka.FULL,
    '0.5':  TeacherLoad.Stavka.HALF,
    '0.25': TeacherLoad.Stavka.QUARTER,
}


def _detect_month(sheet_name: str) -> int | None:
    """
    Sheet nomidan oy raqamini aniqlash.

    HAQIQIY BUG (topilgan): avval faqat aniq mos matn qabul qilinardi
    (masalan "Yanvar") — "Yanvar 2026", "1-Yanvar", "Yanvar oyi" kabi ozgina
    farq qiladigan sheet nomlari umuman tanilmay, o'sha oy butunlay
    o'tkazib yuborilardi (hech qanday ogohlantirishsiz — "N ta oy yuklandi"
    xabari faqat muvaffaqiyatli aniqlanganlarni sanardi).

    Endi: sheet nomi so'zlarga (harf bo'lmagan belgilar bo'yicha) bo'linadi va
    har bir so'z alohida oy nomiga solishtiriladi — shunda oy nomi atrofida
    raqam/sana/qo'shimcha so'z bo'lsa ham to'g'ri aniqlanadi.
    """
    name = sheet_name.strip().lower()
    if name in MONTH_MAP:
        return MONTH_MAP[name]
    for word in re.split(r'[^a-zʻʼ]+', name):
        if word in MONTH_MAP:
            return MONTH_MAP[word]
    return None


def _normalize_stavka(val) -> str:
    """Excel stavka qiymatini modeldagi choicega o'tkazish."""
    if pd.isna(val):
        return TeacherLoad.Stavka.VACANT
    s = str(val).strip().lower()
    if 'vokant' in s or 'vakant' in s:
        return TeacherLoad.Stavka.VACANT
    if 'soatbay' in s or 'soat' in s:
        return TeacherLoad.Stavka.HOURLY
    return STAVKA_MAP.get(s, TeacherLoad.Stavka.FULL)


def _try_find_teacher(full_name: str, organization) -> Teacher | None:
    """F.I.Sh bo'yicha o'qituvchini izlash."""
    if not full_name or not full_name.strip():
        return None
    parts = full_name.strip().split()
    qs = Teacher.objects.filter(organization=organization)
    for part in parts:
        if len(part) > 2:
            qs = qs.filter(user__last_name__icontains=part)
            if qs.count() == 1:
                return qs.first()
    return None


def _try_find_group(group_name: str, organization, month: int, year: int) -> Group | None:
    """Guruh nomiga mos Group topish."""
    if not group_name:
        return None
    return Group.objects.filter(
        organization=organization,
        name__icontains=group_name.strip(),
        month=month,
        year=year,
    ).first()


def _try_find_subject(module_name: str, department: Department) -> CurriculumSubject | None:
    """Modul nomi bo'yicha CurriculumSubject topish (kodni ajratib)."""
    if not module_name:
        return None
    # "3.5. Sport mashg'uloti..." → "3.5" kodini ajrat
    parts = module_name.strip().split('.')
    if len(parts) >= 2:
        code_part = '.'.join(parts[:2]).strip().rstrip('.')
        qs = CurriculumSubject.objects.filter(
            block__department=department,
            subject__code__icontains=code_part,
        )
        if qs.count() == 1:
            return qs.first()
    # Fallback: subject nomi bo'yicha qidirish
    name_part = module_name.split('.')[-1].strip()[:50]
    return CurriculumSubject.objects.filter(
        block__department=department,
        subject__name__icontains=name_part,
    ).first()


def _safe_iloc(row, idx, default=None):
    """iloc xavfsiz versiyasi — ustun mavjud bo'lmasa default qaytaradi."""
    try:
        val = row.iloc[idx]
        return val if pd.notna(val) else default
    except IndexError:
        return default


def parse_load_sheet_excel(file, department: Department, year: int,
                            organization, uploaded_by) -> dict:
    """
    Taqsimot Excel faylini o'qib, LoadSheet + TeacherLoad + LoadDistribution
    larni yaratadi. Har bir sheet = bir oy.
    Qayta yuklansa, avvalgi ma'lumotlar o'chiriladi.

    Excel format (0-indeksdan):
      Qator 0: sarlavha (ixtiyoriy)
      Qator 1: ustun nomlari — T/r | Modul | Guruh1..N | Jami | Hammasi | Stavka | F.I.Sh | Lavozim
      Qator 2+: ma'lumot qatorlari
    Ustunlar:
      0=T/r, 1=Modul, 2..22=Guruhlar (max 21 ta),
      23=Jami, 24=Hammasi, 25=Stavka, 26=F.I.Sh, 27=Lavozim
    """
    xl = pd.ExcelFile(file)
    results = []
    # O'tkazib yuborilgan sheet'lar — foydalanuvchiga sababi bilan ko'rsatish uchun
    # (haqiqiy bug: avval bu haqda hech qanday xabar qaytmasdi, oy "yo'qolib" ketardi)
    skipped = []

    for sheet_name in xl.sheet_names:
        month = _detect_month(sheet_name)
        if month is None:
            # 'Yopilgan guruhlar' kabi haqiqatan oy bo'lmagan sheet'lar ham,
            # nomi tanilmagan haqiqiy oy sheet'lari ham shu yerga tushadi —
            # foydalanuvchi ro'yxatni ko'rib, kerak bo'lsa sheet nomini to'g'rilaydi
            skipped.append(f"«{sheet_name}» — oy nomi aniqlanmadi")
            continue

        df = pd.read_excel(xl, sheet_name=sheet_name, header=None)

        if len(df) < 2:
            skipped.append(f"«{sheet_name}» — bo'sh (ma'lumot qatorlari yo'q)")
            continue  # Bo'sh yoki faqat sarlavha bor sheet

        ncols = len(df.columns)

        # Sarlavha qatori (1-indeks) — guruh nomlari ustun bo'yicha
        header_row = df.iloc[1]
        # Guruh ustunlari: 2 dan 22 gacha (23=Jami, 24=Hammasi, 25=Stavka, 26=FISh, 27=Lavozim)
        group_cols = {}  # {col_index: group_name}
        for col_idx in range(2, min(23, ncols)):
            val = _safe_iloc(header_row, col_idx)
            if val is not None and str(val).strip():
                group_cols[col_idx] = str(val).strip()

        # Avvalgi LoadSheet ni o'chirish (qayta yuklash)
        LoadSheet.objects.filter(
            department=department, month=month, year=year
        ).delete()

        with transaction.atomic():
            load_sheet = LoadSheet.objects.create(
                department=department,
                month=month,
                year=year,
                uploaded_by=uploaded_by,
            )

            current_teacher_load = None

            for row_idx in range(2, len(df)):
                row = df.iloc[row_idx]

                tr_val    = _safe_iloc(row, 0)         # T/r
                modul_val = _safe_iloc(row, 1)         # Modul nomi
                jami_val  = _safe_iloc(row, 23, 0)     # Jami
                hammasi   = _safe_iloc(row, 24, 0)     # Hammasi
                stavka    = _safe_iloc(row, 25)        # Stavka
                fish      = _safe_iloc(row, 26, '')    # F.I.Sh
                lavozim   = _safe_iloc(row, 27, '')    # Lavozim

                # Modul nomi yo'q → oxirgi qator (izoh yoki bo'sh)
                if modul_val is None or not str(modul_val).strip():
                    continue

                module_name = str(modul_val).strip()
                try:
                    row_hours = int(jami_val) if jami_val else 0
                except (ValueError, TypeError):
                    row_hours = 0

                # Yangi o'qituvchi boshlandi (T/r soni bor)
                if tr_val is not None and str(tr_val).strip().isdigit():
                    full_name = str(fish).strip()
                    position  = str(lavozim).strip()
                    stavka_v  = _normalize_stavka(stavka)
                    try:
                        total_h = int(hammasi) if hammasi else 0
                    except (ValueError, TypeError):
                        total_h = 0

                    teacher = _try_find_teacher(full_name, organization)

                    current_teacher_load = TeacherLoad.objects.create(
                        load_sheet=load_sheet,
                        teacher=teacher,
                        full_name=full_name,
                        position=position,
                        stavka=stavka_v,
                        total_hours=total_h,
                    )

                # current_teacher_load yo'q bo'lsa o'tkazib yuborish
                if current_teacher_load is None or row_hours == 0:
                    continue

                # Guruh bo'yicha taqsimot yozuvlari
                for col_idx, g_name in group_cols.items():
                    cell_val = row.iloc[col_idx]
                    if pd.isna(cell_val) or cell_val == 0:
                        continue
                    hours = int(cell_val)
                    if hours <= 0:
                        continue

                    group   = _try_find_group(g_name, organization, month, year)
                    subject = _try_find_subject(module_name, department)

                    LoadDistribution.objects.create(
                        teacher_load=current_teacher_load,
                        curriculum_subject=subject,
                        module_name=module_name,
                        group=group,
                        group_name=g_name,
                        hours=hours,
                    )

        results.append({
            'sheet':    sheet_name,
            'month':    month,
            'year':     year,
            'teachers': load_sheet.teacher_loads.count(),
            'entries':  LoadDistribution.objects.filter(
                            teacher_load__load_sheet=load_sheet).count(),
        })

    return {'sheets': results, 'skipped': skipped}


class TeacherPagination(PageNumberPagination):
    """
    Global PAGE_SIZE=20 dan farqli — ?page_size= orqali oshirish imkonini beradi.
    LoadSheetPage.jsx (`getTeachers({ page_size: 500 })`, "Boshqalar" dropdown guruhi)
    kabi joylar barcha o'qituvchilarni bitta so'rovda olishga tayanadi;
    page_size_query_param sozlanmagan bo'lsa bu parametr jimgina e'tiborga olinmay,
    doim faqat birinchi 20 tasi qaytardi (haqiqiy bug — group-day-assignments.md
    dagi pagination-trap bilan bir xil turkum).
    """
    page_size_query_param = 'page_size'
    max_page_size = 500


class TeacherViewSet(viewsets.ModelViewSet):
    serializer_class = TeacherSerializer
    # O'qituvchilar ro'yxati (o'qish) bir nechta sahifada dropdown/filtr sifatida
    # kerak (SchedulePage.jsx, LoadSheetPage.jsx, TeacherSubjectAssignPage.jsx) —
    # `edu_admin` ham shulardan foydalanadi (`/schedules` route'i unga ochiq),
    # lekin avvalgi `IsDeptManager` faqat super_admin/org_admin/dept_manager'ga
    # ruxsat berardi, natijada edu_admin doim 403 olardi (haqiqiy bug, tuzatilgan).
    # Yozish (yaratish/o'chirish) hamon faqat dept_manager+ bilan cheklangan.
    permission_classes = [IsDeptManagerOrReadOnly]
    pagination_class = TeacherPagination

    def get_queryset(self):
        qs = Teacher.objects.filter(
            organization=self.request.user.organization,
            is_active=True
        ).select_related('user', 'personal_room', 'department').annotate(
            # Fan biriktirish (TeacherSubjectAssignment) bo'yicha sanoqlar
            assigned_major_count=Count('subject_assignments', distinct=True),
            assigned_subject_count=Count('subject_assignments__subjects', distinct=True),
        )

        user = self.request.user
        if user.role == 'dept_manager':
            dept = Department.objects.filter(
                manager=user, organization=user.organization
            ).first()
            if dept:
                qs = qs.filter(department=dept)

        return qs


class TeacherBusyTimePagination(PageNumberPagination):
    """
    Global PAGE_SIZE=20 dan farqli — ?page_size= orqali oshirish imkonini beradi.
    BusyTimesPage.jsx'dagi kalendar (bitta o'qituvchi + bitta oy) 31 tagacha yozuv
    qaytarishi mumkin — page_size_query_param sozlanmagan bo'lsa bu jimgina e'tiborga
    olinmay, faqat birinchi 20 tasi qaytardi (xuddi Group/Room'da uchragan bug bilan
    bir xil turkum — group-day-assignments.md).
    """
    page_size_query_param = 'page_size'
    max_page_size = 500


class TeacherBusyTimeViewSet(viewsets.ModelViewSet):
    """
    O'qituvchining band vaqtlarini boshqarish.
    Kafedra mudiri o'z kafedrasi o'qituvchilari uchun kiritadi.

    Filtrlar (query params):
      ?teacher_id=X     — faqat bitta o'qituvchi
      ?date_from=YYYY-MM-DD
      ?date_to=YYYY-MM-DD
      ?month=M&year=YYYY

    Maxsus actionlar:
      POST /bulk-create/  — bir vaqtda ko'p sana uchun band vaqt qo'shish
      GET  /by-teacher/{teacher_id}/  — o'qituvchi band vaqtlari
    """
    serializer_class   = TeacherBusyTimeSerializer
    permission_classes = [IsDeptManager]
    pagination_class   = TeacherBusyTimePagination

    def get_queryset(self):
        qs     = TeacherBusyTime.objects.filter(
            teacher__organization=self.request.user.organization
        ).select_related('teacher__user')

        # Faqat o'z kafedrasi o'qituvchilari (dept_manager)
        user = self.request.user
        if user.role == 'dept_manager':
            dept = Department.objects.filter(
                manager=user, organization=user.organization
            ).first()
            if dept:
                qs = qs.filter(teacher__department=dept)

        # Query param filtrlari
        params = self.request.query_params
        if teacher_id := params.get('teacher_id'):
            qs = qs.filter(teacher_id=teacher_id)
        if date_from := params.get('date_from'):
            qs = qs.filter(date__gte=date_from)
        if date_to := params.get('date_to'):
            qs = qs.filter(date__lte=date_to)
        if month := params.get('month'):
            qs = qs.filter(date__month=month)
        if year := params.get('year'):
            qs = qs.filter(date__year=year)

        return qs.order_by('date', 'teacher', 'start_time')

    @action(detail=False, methods=['post'], url_path='bulk-create')
    def bulk_create(self, request):
        """
        Bir o'qituvchi uchun bir necha sana/vaqtni bir yo'la qo'shish.

        Body:
        {
          "teacher_id": 5,
          "is_all_day": true,         ← ixtiyoriy (default: false)
          "start_time": "09:00",      ← is_all_day=false bo'lsa majburiy
          "end_time": "11:00",
          "reason": "Konferentsiya",
          "dates": ["2026-04-07", "2026-04-08", "2026-04-09"]
        }
        """
        teacher_id = request.data.get('teacher_id')
        dates      = request.data.get('dates', [])
        is_all_day = request.data.get('is_all_day', False)
        start_time = request.data.get('start_time')
        end_time   = request.data.get('end_time')
        reason     = request.data.get('reason', '')

        if not teacher_id:
            return Response(
                {'error': 'teacher_id majburiy!'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if not dates:
            return Response(
                {'error': 'dates ro\'yxati bo\'sh!'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if not is_all_day and (not start_time or not end_time):
            return Response(
                {'error': 'is_all_day=False bo\'lsa start_time va end_time majburiy!'},
                status=status.HTTP_400_BAD_REQUEST
            )

        teacher = Teacher.objects.filter(
            id=teacher_id,
            organization=request.user.organization
        ).first()
        if not teacher:
            return Response(
                {'error': 'O\'qituvchi topilmadi!'},
                status=status.HTTP_404_NOT_FOUND
            )

        created  = []
        skipped  = []
        errors   = []

        for date_str in dates:
            try:
                date_val = datetime.date.fromisoformat(str(date_str).strip())
            except ValueError:
                errors.append({'date': date_str, 'error': 'Noto\'g\'ri sana formati (YYYY-MM-DD)'})
                continue

            # Serializer orqali validatsiya
            data = {
                'teacher':    teacher.id,
                'date':       date_val,
                'is_all_day': is_all_day,
                'start_time': start_time if not is_all_day else None,
                'end_time':   end_time   if not is_all_day else None,
                'reason':     reason,
            }
            ser = TeacherBusyTimeSerializer(data=data)
            if ser.is_valid():
                obj = ser.save()
                created.append(TeacherBusyTimeSerializer(obj).data)
            else:
                # Agar qoplanish bo'lsa — skip
                skipped.append({'date': date_str, 'error': ser.errors})

        return Response({
            'created': created,
            'skipped': skipped,
            'errors':  errors,
        }, status=status.HTTP_201_CREATED if created else status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'],
            url_path=r'by-teacher/(?P<teacher_id>[0-9]+)')
    def by_teacher(self, request, teacher_id=None):
        """
        GET /teacher-busy-times/by-teacher/{teacher_id}/
        O'qituvchining band vaqtlari + xulosa.

        Query params: ?month=M&year=YYYY
        """
        qs = self.get_queryset().filter(teacher_id=teacher_id)
        params = request.query_params
        if month := params.get('month'):
            qs = qs.filter(date__month=month)
        if year := params.get('year'):
            qs = qs.filter(date__year=year)

        data = TeacherBusyTimeSerializer(qs, many=True).data
        return Response({
            'teacher_id':  teacher_id,
            'total_count': len(data),
            'busy_times':  data,
        })


class TeacherSubjectAssignmentViewSet(viewsets.ModelViewSet):
    serializer_class = TeacherSubjectAssignmentSerializer
    permission_classes = [IsDeptManager]

    def get_queryset(self):
        qs = TeacherSubjectAssignment.objects.filter(
            teacher__organization=self.request.user.organization
        ).select_related('teacher__user', 'major').prefetch_related('subjects')

        user = self.request.user
        if user.role == 'dept_manager':
            dept = Department.objects.filter(
                manager=user, organization=user.organization
            ).first()
            if dept:
                qs = qs.filter(teacher__department=dept)

        major_id = self.request.query_params.get('major_id')
        if major_id:
            qs = qs.filter(major_id=major_id)
        teacher_id = self.request.query_params.get('teacher_id')
        if teacher_id:
            qs = qs.filter(teacher_id=teacher_id)
        return qs

    @action(detail=False, methods=['post'], url_path='bulk-assign')
    def bulk_assign(self, request):
        """
        POST /api/v1/teacher-subject-assignments/bulk-assign/
        Bir yo'la ko'p o'qituvchiga fanlar biriktirish (yoki olib tashlash).

        Body:
        {
          "major_id": 1,
          "assignments": [
            {"teacher_id": 3, "subject_ids": [1, 2, 5]},
            {"teacher_id": 4, "subject_ids": [2, 3]},
            {"teacher_id": 5, "subject_ids": []}     ← bo'sh = biriktirishni olib tashlash
          ]
        }
        """
        from academic.models import Major, Subject as AcSubject

        major_id    = request.data.get('major_id')
        assignments = request.data.get('assignments', [])
        org         = request.user.organization

        if not major_id:
            return Response({'error': 'major_id majburiy'}, status=status.HTTP_400_BAD_REQUEST)

        major = Major.objects.filter(id=major_id, organization=org).first()
        if not major:
            return Response({'error': 'Yo\'nalish topilmadi'}, status=status.HTTP_404_NOT_FOUND)

        # Dept manager faqat o'z kafedrasi o'qituvchilariga biriktira oladi
        dept = None
        if request.user.role == 'dept_manager':
            dept = Department.objects.filter(
                manager=request.user, organization=org
            ).first()

        results = []
        with transaction.atomic():
            for item in assignments:
                teacher_id  = item.get('teacher_id')
                subject_ids = item.get('subject_ids', [])

                teacher_qs = Teacher.objects.filter(id=teacher_id, organization=org)
                if dept:
                    teacher_qs = teacher_qs.filter(department=dept)
                teacher = teacher_qs.first()
                if not teacher:
                    continue

                if not subject_ids:
                    # Bo'sh — mavjud biriktirishni o'chirish
                    TeacherSubjectAssignment.objects.filter(
                        teacher=teacher, major=major
                    ).delete()
                    results.append({
                        'teacher_id': teacher_id,
                        'action': 'removed',
                        'subjects': 0,
                    })
                    continue

                subjects = AcSubject.objects.filter(
                    id__in=subject_ids, organization=org
                )
                assignment, created = TeacherSubjectAssignment.objects.get_or_create(
                    teacher=teacher, major=major
                )
                assignment.subjects.set(subjects)
                results.append({
                    'teacher_id':  teacher_id,
                    'teacher_name': teacher.user.get_full_name(),
                    'action':      'created' if created else 'updated',
                    'subjects':    subjects.count(),
                })

        return Response({
            'success': True,
            'message': f'{len(results)} ta o\'qituvchi uchun fan biriktiruvi yangilandi.',
            'results': results,
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='majors-subjects',
            permission_classes=[IsAuthenticated])
    def majors_subjects(self, request):
        """
        GET /api/v1/teacher-subject-assignments/majors-subjects/?teacher_id=X

        Yo'nalishlar (Major) va faol o'quv rejasidagi fanlari.
        teacher_id berilsa — faqat o'qituvchining KAFEDRASIGA tegishli fanlar
        (CurriculumSubject.department == teacher.department) qaytariladi va
        faqat shunday fani bor yo'nalishlar ko'rsatiladi. O'qituvchida kafedra
        belgilanmagan bo'lsa fan qaytmaydi.

        Javob: {"teacher_has_department": bool, "majors": [{major_id, major_name, subjects: [...]}]}
        """
        from academic.models import Major, Curriculum

        org = request.user.organization

        teacher_id = request.query_params.get('teacher_id')
        dept_id = None
        if teacher_id:
            teacher = Teacher.objects.filter(
                id=teacher_id, organization=org
            ).first()
            if not teacher or not teacher.department_id:
                # Kafedra belgilanmagan — fan yo'q
                return Response({'teacher_has_department': False, 'majors': []})
            dept_id = teacher.department_id

        majors = Major.objects.filter(organization=org, is_active=True).order_by('name')

        result = []
        today = datetime.date.today()
        for major in majors:
            # O'qituvchini fanga biriktirish uning onlayn/oflayn dars berish-bermasligiga
            # bog'liq emas — shu yo'nalishning HAR IKKI turdagi (oflayn va onlayn) faol
            # o'quv rejasidan fanlar birlashtiriladi (bir xil fan ikkalasida ham bo'lishi
            # mumkin, seen_ids orqali dublikat qilinmaydi)
            subjects = []
            seen_ids = set()
            for dm in (DeliveryMode.OFFLINE, DeliveryMode.ONLINE):
                curriculum = Curriculum.get_active_for_date(
                    major, target_date=today, delivery_mode=dm,
                    queryset=Curriculum.objects.prefetch_related('blocks__subjects__subject'),
                )
                if not curriculum:
                    continue
                for block in curriculum.blocks.all():
                    for cs in block.subjects.all():
                        # Kafedra bo'yicha filtr (teacher_id berilgan bo'lsa)
                        if dept_id is not None and cs.department_id != dept_id:
                            continue
                        if cs.subject_id in seen_ids:
                            continue
                        seen_ids.add(cs.subject_id)
                        subjects.append({
                            'id':   cs.subject.id,
                            'name': cs.subject.name,
                        })

            # teacher_id berilganda — faqat fani bor yo'nalishlarni ko'rsatamiz
            if dept_id is not None and not subjects:
                continue

            result.append({
                'major_id':   major.id,
                'major_name': major.name,
                'subjects':   subjects,
            })

        return Response({'teacher_has_department': True, 'majors': result})


class TeacherMonthlyLoadViewSet(viewsets.ModelViewSet):
    serializer_class = TeacherMonthlyLoadSerializer
    permission_classes = [IsDeptManager]

    def get_queryset(self):
        return TeacherMonthlyLoad.objects.filter(
            teacher__organization=self.request.user.organization
        ).select_related('teacher__user', 'major')

    def perform_create(self, serializer):
        serializer.save(assigned_by=self.request.user)

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        load = self.get_object()
        load.status = TeacherMonthlyLoad.Status.APPROVED
        load.save()
        return Response({'message': 'Tasdiqlandi'})


class ScheduleViewSet(viewsets.ModelViewSet):
    serializer_class = ScheduleSerializer
    permission_classes = [IsEduAdmin]

    def get_permissions(self):
        # list/retrieve — kafedra mudiri ham "Ko'rish" sahifasida jadvalni
        # (hatto hali `draft` holatida, nashr etilmagan bo'lsa ham) ko'ra olishi
        # kerak — boshqa detail action'lar (by-group/filters/teacher-days/
        # free-teachers) allaqachon shu ViewSet ichida IsAuthenticated bilan
        # ochiq edi, faqat asosiy list/retrieve IsEduAdmin bilan cheklangan edi.
        # Yaratish/generatsiya/nashr/o'chirish hamon faqat IsEduAdmin+.
        if self.action in ('list', 'retrieve'):
            return [IsAuthenticated()]
        return [IsEduAdmin()]

    def get_queryset(self):
        return Schedule.objects.filter(
            organization=self.request.user.organization
        )

    @action(detail=False, methods=['post'], url_path='generate')
    def generate(self, request):
        """
        POST /api/v1/schedules/generate/
        Body: { month, year, title, date_from(ixtiyoriy), date_to(ixtiyoriy), time_limit(ixtiyoriy) }

        Ketma-ketlik:
          1. GroupSubject (haqiqiy Taqsimot manbai) dan vazifalar olinadi
          2. OR-Tools CP-SAT yordamida jadval tuziladi
          3. ScheduleEntry lar yaratiladi
        """
        from .solver import generate_schedule

        month      = request.data.get('month')
        year       = request.data.get('year')
        title      = request.data.get('title')
        org        = request.user.organization
        # ── VAQT LIMITI: pastdan CHEKLANGAN ──────────────────────────────────
        # **Haqiqiy muammo (o'lchangan)**: limit kichik bo'lsa jadval sifati
        # keskin tushadi va buni foydalanuvchi sezmaydi — natija "shunchaki
        # yomonroq" bo'lib chiqadi:
        #   60s  → 1280/1283 joylashadi, 24 o'qituvchidan 16 tasida bo'sh kun
        #   120s → 1282/1283, 18 tasida bo'sh kun
        #   240s → 1283/1283, 0–2 tasida bo'sh kun
        # Sabab: 1-bosqich (maksimal dars soni) o'zi 30–120 soniya talab
        # qiladi; undan kam vaqtda viloyat ixchamligiga umuman navbat
        # yetmaydi. Shuning uchun quyi chegara qat'iy qo'yilgan.
        MIN_TIME_LIMIT = 240
        requested_limit = int(request.data.get('time_limit', MIN_TIME_LIMIT))
        time_limit = max(MIN_TIME_LIMIT, requested_limit)

        if not all([month, year, title]):
            return Response(
                {'error': 'month, year, title majburiy!'},
                status=status.HTTP_400_BAD_REQUEST
            )

        month = int(month)
        year  = int(year)

        # Oy sanalarini hisoblash
        date_from_str = request.data.get('date_from')
        date_to_str   = request.data.get('date_to')

        if date_from_str:
            date_from = datetime.date.fromisoformat(date_from_str)
        else:
            date_from = datetime.date(year, month, 1)

        if date_to_str:
            date_to = datetime.date.fromisoformat(date_to_str)
        else:
            if month == 12:
                date_to = datetime.date(year + 1, 1, 1) - datetime.timedelta(days=1)
            else:
                date_to = datetime.date(year, month + 1, 1) - datetime.timedelta(days=1)

        # Guruh kunlik biriktiruvi (GroupDayAssignment) mavjudligini tekshirish —
        # bu haqiqiy Taqsimot manbai (GroupSubject) uchun zarur old shart, xuddi
        # `curriculum_preview` bilan bir xil tekshiruv (LoadDistribution/Excel emas —
        # batafsil `.claude/rules/schedule-generation.md`)
        from academic.models import GroupDayAssignment
        if not GroupDayAssignment.objects.filter(
            group__organization=org, date__year=year, date__month=month,
        ).exists():
            return Response(
                {'error': f'{month}/{year} uchun guruh kunlik biriktiruvi topilmadi. '
                          "Avval 'Guruh biriktiruv' bo'limida kunlarga smena va bino "
                          'biriktiring, so\'ng Taqsimotda fanlarga o\'qituvchi tayinlang.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # -- JADVALNI YARATISH/TOZALASH (tez, so'rov ichida) -----------------
        with transaction.atomic():
            schedule, created = Schedule.objects.get_or_create(
                organization=org, month=month, year=year,
                defaults={
                    'title':        title,
                    'date_from':    date_from,
                    'date_to':      date_to,
                    'generated_by': request.user,
                }
            )
            # "Osilib qolgan" jarayon qayta generatsiyani abadiy bloklab
            # qo'ymasligi uchun eskirganini tekshiramiz (server qayta ishga
            # tushsa oqim o'ladi, holat esa `running` bo'lib qolaveradi).
            if (schedule.gen_status == Schedule.GenStatus.RUNNING
                    and schedule.gen_started
                    and (timezone.now() - schedule.gen_started).total_seconds()
                        > STALE_GEN_SECONDS):
                schedule.gen_status = Schedule.GenStatus.FAILED
                schedule.save(update_fields=['gen_status'])

            if schedule.gen_status == Schedule.GenStatus.RUNNING:
                return Response(
                    {'error': 'Bu oy uchun jadval hozir tuzilmoqda. '
                              'Jarayon tugashini kuting.',
                     'schedule_id': schedule.id},
                    status=status.HTTP_409_CONFLICT
                )
            if not created:
                # Qayta generatsiya - avvalgi yozuvlar o'chiriladi
                schedule.entries.all().delete()
                schedule.status = Schedule.Status.DRAFT
                schedule.title = title
                schedule.date_from = date_from
                schedule.date_to = date_to
                schedule.generated_by = request.user
            schedule.gen_status = Schedule.GenStatus.RUNNING
            schedule.gen_percent = 0
            schedule.gen_step = 'Jarayon boshlanmoqda'
            schedule.gen_detail = ("Jadval tuzish boshlandi. Bu odatda 4-6 daqiqa "
                                   "davom etadi - oynani yopib ketsangiz ham jarayon "
                                   "davom etaveradi.")
            schedule.gen_started = timezone.now()
            schedule.gen_finished = None
            schedule.gen_result = None
            schedule.save()

        # -- SOLVERNI FON OQIMIDA ISHGA TUSHIRISH ----------------------------
        # **Nega thread**: jadval tuzish 4-6 daqiqa davom etadi. Sinxron
        # so'rovda brauzer/proksi timeout bo'lib ketardi va foydalanuvchi
        # jarayon qay bosqichda ekanini umuman ko'rmasdi. Celery kabi navbat
        # tizimi loyihada yo'q (broker sozlanmagan), shuning uchun eng sodda
        # yechim - alohida oqim + holatni bazaga yozish. Frontend
        # `GET /schedules/{id}/progress/` orqali so'rab turadi, shuning uchun
        # sahifa yopilib qayta ochilsa ham jarayon ko'rinib turadi.
        sch_id = schedule.id
        org_id = org.id

        def _run():
            from organizations.models import Organization as _Org
            try:
                sch = Schedule.objects.get(pk=sch_id)
                _org = _Org.objects.get(pk=org_id)

                def _cb(pct, step, detail=''):
                    Schedule.objects.filter(pk=sch_id).update(
                        gen_percent=pct, gen_step=step, gen_detail=detail)

                result = generate_schedule(
                    schedule=sch, organization=_org, month=month, year=year,
                    time_limit_seconds=time_limit, progress_cb=_cb,
                )
                entries  = result['entries']
                stats    = result['stats']
                warnings = result['warnings']
                if requested_limit < MIN_TIME_LIMIT:
                    warnings.insert(0, (
                        f"Vaqt limiti {requested_limit}s juda kichik edi - "
                        f"{MIN_TIME_LIMIT}s ga oshirildi. Kichik limitda darslar "
                        "to'liq joylashmaydi va o'qituvchilarda bo'sh kunlar "
                        "paydo bo'ladi."
                    ))
                with transaction.atomic():
                    if entries:
                        ScheduleEntry.objects.bulk_create(entries)
                    Schedule.objects.filter(pk=sch_id).update(
                        gen_status=Schedule.GenStatus.DONE,
                        gen_percent=100,
                        gen_step='Jadval tayyor',
                        gen_detail=(
                            f"{stats.get('placed_paras', 0)} ta dars "
                            f"{stats.get('total_paras', 0)} tadan joylashtirildi."),
                        gen_finished=timezone.now(),
                        gen_result={'stats': stats, 'warnings': warnings,
                                    'created': len(entries)},
                    )
            except Exception as e:
                logger.exception("Jadval generatsiyasida kutilmagan xato "
                                 "(schedule=%s, %s/%s)", sch_id, month, year)
                Schedule.objects.filter(pk=sch_id).update(
                    gen_status=Schedule.GenStatus.FAILED,
                    gen_step='Xatolik yuz berdi',
                    gen_detail=str(e),
                    gen_finished=timezone.now(),
                    gen_result={'error': str(e)},
                )
            finally:
                # Oqim o'z DB ulanishini yopishi SHART - aks holda ulanishlar
                # to'planib qoladi (Django har oqim uchun alohida ulanish ochadi).
                db_connection.close()

        threading.Thread(target=_run, daemon=True,
                         name=f'schedule-gen-{sch_id}').start()

        return Response(
            {'schedule': ScheduleSerializer(schedule).data,
             'message': 'Jadval tuzish boshlandi.'},
            status=status.HTTP_202_ACCEPTED
        )

    @action(detail=True, methods=['get'], url_path='progress',
            permission_classes=[IsAuthenticated])
    def progress(self, request, pk=None):
        """
        GET /api/v1/schedules/{id}/progress/

        Generatsiya jarayonining joriy holati - progress bar uchun.
        Sahifa yopilib qayta ochilsa ham shu endpoint orqali jarayon
        ko'rinib turadi (holat bazada saqlanadi, xotirada emas).
        """
        sch = self.get_object()

        # -- "Osilib qolgan" jarayonni aniqlash ------------------------------
        # Server qayta ishga tushsa, oqim o'lib ketadi-yu, holat `running`
        # bo'lib qolaveradi. Bir soatdan ko'p o'tgan bo'lsa - xato deb
        # belgilanadi, aks holda progress bar abadiy aylanib turardi.
        if (sch.gen_status == Schedule.GenStatus.RUNNING and sch.gen_started
                and (timezone.now() - sch.gen_started).total_seconds()
                    > STALE_GEN_SECONDS):
            sch.gen_status = Schedule.GenStatus.FAILED
            sch.gen_step = 'Jarayon uzilib qoldi'
            sch.gen_detail = ("Server qayta ishga tushgan bo'lishi mumkin. "
                              "Jadvalni qaytadan tuzing.")
            sch.gen_finished = timezone.now()
            sch.save()

        elapsed = None
        if sch.gen_started:
            end = sch.gen_finished or timezone.now()
            elapsed = int((end - sch.gen_started).total_seconds())

        return Response({
            'schedule_id': sch.id,
            'title':       sch.title,
            'status':      sch.gen_status,
            'percent':     sch.gen_percent,
            'step':        sch.gen_step,
            'detail':      sch.gen_detail,
            'elapsed_s':   elapsed,
            'entries':     sch.entries.count(),
            'result':      sch.gen_result,
        })

    @action(detail=True, methods=['post'])
    def publish(self, request, pk=None):
        schedule = self.get_object()
        schedule.status = Schedule.Status.PUBLISHED
        schedule.save()
        return Response({'message': 'Jadval nashr etildi'})

    @action(detail=True, methods=['get'], url_path='by-group/(?P<group_id>[0-9]+)',
            permission_classes=[IsAuthenticated])
    def by_group(self, request, pk=None, group_id=None):
        schedule = self.get_object()
        entries  = schedule.entries.filter(
            group_id=group_id
        ).order_by('date', 'para__order')
        return Response(ScheduleEntrySerializer(entries, many=True).data)

    @action(detail=True, methods=['get'], url_path='by-teacher/(?P<teacher_id>[0-9]+)',
            permission_classes=[IsAuthenticated])
    def by_teacher(self, request, pk=None, teacher_id=None):
        schedule = self.get_object()
        entries  = schedule.entries.filter(
            teacher_id=teacher_id
        ).order_by('date', 'para__order')
        return Response(ScheduleEntrySerializer(entries, many=True).data)

    @action(detail=True, methods=['get'], url_path='filters', permission_classes=[IsAuthenticated])
    def filters(self, request, pk=None):
        """
        GET /schedules/{id}/filters/

        Frontend'dagi "Ko'rish" oynasidagi Guruh/O'qituvchi Select'lari uchun —
        **butun tashkilotdagi barcha** guruh/o'qituvchi emas, faqat shu jadvalda
        haqiqatan darsi bor guruh/o'qituvchilar ro'yxatini qaytaradi (haqiqiy bug,
        tuzatilgan: avval frontend `getGroups`/`getTeachers` orqali butun tashkilot
        ro'yxatini olardi, natijada bu jadvalga aloqasi yo'q guruhlar ham dropdown'da
        ko'rinardi).
        """
        # MUHIM: ScheduleEntry.Meta.ordering = ['date', 'para__order'] — agar
        # `.values().distinct()` zanjirida aniq `.order_by(...)` bo'lmasa, Django
        # shu tartiblash maydonlarini so'rovga yashirincha qo'shib qo'yadi va
        # DISTINCT haqiqiy dublikatlarni olib tashlay olmay qoladi (har bir
        # sana/para juftligi "boshqacha qator" deb hisoblanadi). Shuning uchun
        # ikkalasida ham aniq `.order_by(...)` bilan standart tartiblash
        # bekor qilinadi (haqiqiy bug, tuzatilgan).
        schedule = self.get_object()
        groups = (
            schedule.entries
            .filter(group__isnull=False)
            .values('group_id', 'group__name')
            .distinct()
            .order_by('group__name')
        )
        teachers = (
            schedule.entries
            .filter(teacher__isnull=False)
            .values('teacher_id', 'teacher__user__first_name', 'teacher__user__last_name')
            .distinct()
            .order_by('teacher__user__first_name', 'teacher__user__last_name')
        )
        return Response({
            'groups': [
                {'id': g['group_id'], 'name': g['group__name']} for g in groups
            ],
            'teachers': [
                {
                    'id': t['teacher_id'],
                    # `full_name` — frontenddagi `getTeacherName()` shu maydonni kutadi
                    'full_name': f"{t['teacher__user__first_name']} {t['teacher__user__last_name']}".strip(),
                }
                for t in teachers
            ],
        })

    @action(detail=True, methods=['get'], url_path='teacher-days',
            permission_classes=[IsAuthenticated])
    def teacher_days(self, request, pk=None):
        """
        GET /schedules/{id}/teacher-days/

        Har bir o'qituvchi BINO KESIMIDA necha kun dars berishini qaytaradi.

        Nima uchun kerak: viloyat binolariga o'qituvchilar komandirovkaga
        boradi, shuning uchun "kim qayerda necha kun turadi" — rejalashtirish
        va xarajat uchun asosiy ko'rsatkich. `days` (kun soni) — asosiy raqam;
        `span` (birinchi–oxirgi kun oralig'i) va `gap_days` (safar ichidagi
        BO'SH kunlar) safar naqadar ixcham ekanini ko'rsatadi.

        Javob:
        {
          "buildings": [{id, name, is_regional}],
          "rows": [{teacher_id, teacher_name, building_id, building_name,
                    is_regional, days, lessons, first_date, last_date,
                    span, gap_days}],
          "totals": [{teacher_id, teacher_name, days, lessons, buildings}]
        }
        """
        from collections import defaultdict
        from organizations.models import Building

        schedule = self.get_object()
        entries = (
            schedule.entries
            .filter(teacher__isnull=False)
            .select_related('teacher__user', 'building')
        )

        # (teacher_id, building_id) -> {sanalar}
        per = defaultdict(set)
        lessons = defaultdict(int)
        names = {}
        bnames = {}
        per_teacher_days = defaultdict(set)
        for en in entries:
            key = (en.teacher_id, en.building_id)
            per[key].add(en.date)
            lessons[key] += 1
            per_teacher_days[en.teacher_id].add(en.date)
            if en.teacher_id not in names:
                u = en.teacher.user
                names[en.teacher_id] = (u.get_full_name() or u.username).strip()
            if en.building_id and en.building_id not in bnames:
                bnames[en.building_id] = en.building.name

        regional = set(
            Building.objects
            .filter(organization=schedule.organization, is_regional=True)
            .values_list('id', flat=True)
        )

        # -- ISH KUNLARI INDEKSI (haqiqiy bug, tuzatilgan) -------------------
        # Avval `span` KALENDAR kunlari bo'yicha hisoblanardi:
        #     span = (oxirgi_sana - birinchi_sana).days + 1
        # Natijada orada qolgan YAKSHANBA ham "bo'sh kun" bo'lib sanalardi.
        # Real misol: 24 o'qituvchidan 11 tasida "1 kun bo'sh" ko'rinardi,
        # aslida hammasi ketma-ket ishlagan - o'sha 1 kun dam olish kuni edi.
        # Solver esa bo'shliqni ISH KUNLARI bo'yicha o'lchaydi, ya'ni hisobot
        # solver optimallashtirayotgan narsadan boshqa raqam ko'rsatardi.
        # Endi indeks jadvalda dars bor kunlar bo'yicha quriladi - hisobot
        # solver bilan bir xil narsani ko'rsatadi.
        work_days = sorted({en.date for en in entries})
        day_ix = {d: i for i, d in enumerate(work_days)}

        rows = []
        for (t_id, b_id), dates in per.items():
            ds = sorted(dates)
            span = (day_ix[ds[-1]] - day_ix[ds[0]] + 1) if ds else 0
            rows.append({
                'teacher_id':    t_id,
                'teacher_name':  names.get(t_id, ''),
                'building_id':   b_id,
                'building_name': bnames.get(b_id) or 'Onlayn (Zoom)',
                'is_regional':   b_id in regional,
                'days':          len(ds),
                'lessons':       lessons[(t_id, b_id)],
                'first_date':    ds[0] if ds else None,
                'last_date':     ds[-1] if ds else None,
                'span':          span,
                # Safar ichidagi bo'sh kunlar — ISH KUNLARI bo'yicha
                # (yakshanba va dars umuman bo'lmagan kunlar hisobga olinmaydi)
                'gap_days':      max(0, span - len(ds)),
            })
        rows.sort(key=lambda r: (-r['is_regional'], -r['days'], r['teacher_name']))

        totals = [
            {
                'teacher_id':   t_id,
                'teacher_name': names.get(t_id, ''),
                'days':         len(dates),
                'lessons':      sum(v for (tt, _b), v in lessons.items() if tt == t_id),
                'buildings':    sum(1 for (tt, _b) in per if tt == t_id),
            }
            for t_id, dates in per_teacher_days.items()
        ]
        totals.sort(key=lambda r: (-r['days'], r['teacher_name']))

        used_b = {r['building_id'] for r in rows if r['building_id']}
        return Response({
            'buildings': [
                {'id': b.id, 'name': b.name, 'is_regional': b.is_regional}
                for b in Building.objects.filter(id__in=used_b).order_by('name')
            ],
            'rows':   rows,
            'totals': totals,
        })

    @action(detail=True, methods=['get'], url_path='free-teachers',
            permission_classes=[IsAuthenticated])
    def free_teachers(self, request, pk=None):
        """
        GET /schedules/{id}/free-teachers/?date_from=YYYY-MM-DD&date_to=YYYY-MM-DD

        Berilgan sana oralig'ida (odatda bitta hafta) har bir kun va har bir
        PARA (aniq vaqt uyachasi) uchun BO'SH o'qituvchilar ro'yxati — guruh
        yoki bino tanlanmaydi, butun tashkilot bo'yicha.

        **Muhim**: paralar `Para.id` bo'yicha emas, HAQIQIY VAQT (start_time,
        end_time) bo'yicha guruhlanadi — turli smenalarning bir xil vaqtga
        to'g'ri keluvchi paralari (masalan ikkita "1-para" 09:00-10:20) bitta
        ustunga birlashadi. Bu aynan solver.py Constraint 2'dagi bilan bir xil
        mantiq (`.claude/rules/schedule-generation.md`) — aks holda bir xil
        vaqt ikki marta ko'rsatilib, chalkashlik keltirib chiqarardi.

        "Bo'sh" — shu jadvalda o'sha (sana, vaqt) da darsi yo'q VA
        `TeacherBusyTime` orqali band deb belgilanmagan o'qituvchi.

        Har bir o'qituvchi KAFEDRASI (`Teacher.department`) bilan birga
        qaytariladi va natija shu bo'yicha guruhlangan holda chiqadi
        (`slot.departments`) — kafedra mudiri faqat o'z kafedrasini,
        edu_admin/org_admin esa barcha kafedralarni kafedra kesimida ko'radi.

        Qo'shimcha `?department=<id>` bilan aniq bitta kafedraga filtrlash
        mumkin. `dept_manager` uchun bu avtomatik qo'llanadi — boshqa
        rollar `roles-permissions.md` dagi bir xil naqsh: `Teacher.department`
        FK orqali scoping (avval `Teacher.subjects` orqali bilvosita scoping
        qilingan va real ma'lumotda umuman ishlamagan edi).
        """
        from collections import defaultdict

        schedule = self.get_object()
        org = schedule.organization

        date_from_str = request.query_params.get('date_from')
        date_to_str   = request.query_params.get('date_to')
        if not date_from_str or not date_to_str:
            return Response(
                {'error': 'date_from va date_to majburiy (YYYY-MM-DD)'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        date_from = datetime.date.fromisoformat(date_from_str)
        date_to   = datetime.date.fromisoformat(date_to_str)

        # Ish kunlari — Yakshanba dars yo'q (solver.py bilan bir xil qoida)
        working_days = []
        d = date_from
        while d <= date_to:
            if d.isoweekday() <= 6:
                working_days.append(d)
            d += datetime.timedelta(days=1)

        # Distinct vaqt uyachalari — barcha smenalarning barcha paralari
        # o'rtasidan, boshlanish vaqti bo'yicha tartiblanib, bir xil vaqtlar
        # bitta ustunga birlashtiriladi
        all_paras = (
            Para.objects.filter(shift__organization=org, is_active=True)
            .order_by('start_time')
        )
        slots, seen = [], set()
        for p in all_paras:
            key = (p.start_time, p.end_time)
            if key in seen:
                continue
            seen.add(key)
            slots.append(p)   # vakillik qiluvchi Para — is_conflict() uchun yetarli

        teachers_qs = (
            Teacher.objects.filter(organization=org, is_active=True)
            .select_related('user', 'department')
        )

        # `dept_manager` avtomatik faqat o'z kafedrasini ko'radi — boshqa
        # ViewSet'lardagi bir xil naqsh (`TeacherViewSet.get_queryset`,
        # roles-permissions.md). Shu bilan birga filtr Select'i uchun
        # qaytariladigan `departments_out` ham SHU kafedra bilan cheklanadi —
        # aks holda dept_manager javobda boshqa kafedralarni ham ko'rardi,
        # garchi tanlash imkoni bo'lmasa ham (chalkashtiruvchi).
        user = request.user
        if user.role == 'dept_manager':
            dept = Department.objects.filter(
                manager=user, organization=org
            ).first()
            if dept:
                teachers_qs = teachers_qs.filter(department=dept)
                departments_out = [{'id': dept.id, 'name': dept.name}]
            else:
                departments_out = list(
                    Department.objects.filter(organization=org)
                    .order_by('order', 'name').values('id', 'name')
                )
        else:
            # Kafedralar ro'yxati (filtr Select'i uchun) — shu tashkilotdagi
            # BARCHA kafedralar, teacher_data'dagi haqiqatan mavjud
            # bo'lganlari emas (aks holda kafedra tanlangach ro'yxat bo'sh
            # chiqib qolardi)
            departments_out = list(
                Department.objects.filter(organization=org)
                .order_by('order', 'name').values('id', 'name')
            )
            if dept_param := request.query_params.get('department'):
                teachers_qs = teachers_qs.filter(department_id=dept_param)

        teachers = list(teachers_qs)
        teacher_data = [
            {
                'id': t.id,
                'full_name': (t.user.get_full_name() or t.user.username).strip()
                             if t.user else f'#{t.id}',
                'department_id':   t.department_id,
                'department_name': t.department.name if t.department else "Kafedrasiz",
            }
            for t in teachers
        ]

        # (sana, start_time, end_time) -> band o'qituvchi id'lar (shu jadvalda)
        busy_by_slot = defaultdict(set)
        for e in (
            ScheduleEntry.objects
            .filter(schedule=schedule, date__range=(date_from, date_to),
                     teacher__isnull=False)
            .select_related('para')
        ):
            busy_by_slot[(e.date, e.para.start_time, e.para.end_time)].add(e.teacher_id)

        # TeacherBusyTime — shu davr uchun bitta so'rovda
        busy_times = list(
            TeacherBusyTime.objects.filter(
                teacher__organization=org, date__range=(date_from, date_to),
            )
        )

        days_out = []
        for day in working_days:
            slot_out = []
            for idx, slot in enumerate(slots):
                busy_ids = set(busy_by_slot.get((day, slot.start_time, slot.end_time), ()))
                for bt in busy_times:
                    if bt.is_conflict(day, slot):
                        busy_ids.add(bt.teacher_id)
                free = [t for t in teacher_data if t['id'] not in busy_ids]

                # Kafedra bo'yicha guruhlash — `Department.order` tartibida,
                # so'ng "Kafedrasiz" (department=null) eng oxirida
                by_dept = defaultdict(list)
                for t in free:
                    by_dept[(t['department_id'], t['department_name'])].append(t)
                dept_order = {d['id']: i for i, d in enumerate(departments_out)}
                dept_out = [
                    {
                        'department_id':   dept_id,
                        'department_name': dept_name,
                        'free_count':      len(dept_teachers),
                        'teachers':        dept_teachers,
                    }
                    for (dept_id, dept_name), dept_teachers in sorted(
                        by_dept.items(),
                        key=lambda kv: (dept_order.get(kv[0][0], 999), kv[0][1]),
                    )
                ]

                slot_out.append({
                    'label':      f'{idx + 1}-para',
                    'start_time': slot.start_time,
                    'end_time':   slot.end_time,
                    'free_count': len(free),
                    'departments': dept_out,
                })
            days_out.append({'date': day, 'slots': slot_out})

        return Response({
            'slot_count': len(slots),
            'total_teachers': len(teacher_data),
            'departments': departments_out,
            'days': days_out,
        })


class ScheduleEntryViewSet(viewsets.ModelViewSet):
    """
    Jadval elementlari (darslar) — alohida CRUD.

    Filtrlar:
      ?schedule_id=X
      ?group_id=X
      ?teacher_id=X
      ?date=YYYY-MM-DD
      ?date_from=YYYY-MM-DD
      ?date_to=YYYY-MM-DD
    """
    serializer_class   = ScheduleEntrySerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = (
            ScheduleEntry.objects
            .filter(schedule__organization=self.request.user.organization)
            .select_related(
                'teacher__user', 'group', 'subject',
                'para', 'room', 'building', 'schedule'
            )
            .order_by('date', 'para__order')
        )
        p = self.request.query_params
        if v := p.get('schedule_id'): qs = qs.filter(schedule_id=v)
        if v := p.get('group_id'):    qs = qs.filter(group_id=v)
        if v := p.get('teacher_id'): qs = qs.filter(teacher_id=v)
        if v := p.get('date'):        qs = qs.filter(date=v)
        if v := p.get('date_from'):   qs = qs.filter(date__gte=v)
        if v := p.get('date_to'):     qs = qs.filter(date__lte=v)
        return qs


class SubstitutionViewSet(viewsets.ModelViewSet):
    serializer_class = SubstitutionSerializer
    permission_classes = [IsDeptManager]

    def get_queryset(self):
        return Substitution.objects.filter(
            schedule_entry__teacher__organization=self.request.user.organization
        )

    def perform_create(self, serializer):
        serializer.save(requested_by=self.request.user)

    @action(detail=False, methods=['post'], url_path='find-available')
    def find_available(self, request):
        """O'sha kuni bo'sh o'qituvchilarni topish"""
        entry_id = request.data.get('schedule_entry_id')
        date     = request.data.get('date')

        try:
            entry = ScheduleEntry.objects.get(id=entry_id)
        except ScheduleEntry.DoesNotExist:
            return Response({'error': 'Topilmadi'}, status=404)

        subject = entry.subject
        para    = entry.para

        # O'sha parada band o'qituvchilar
        busy_ids = ScheduleEntry.objects.filter(
            date=date,
            para=para
        ).values_list('teacher_id', flat=True)

        # Bo'sh va fanni o'qita oladigan o'qituvchilar
        available = Teacher.objects.filter(
            organization=self.request.user.organization,
            subject_assignments__subjects=subject,
            is_active=True
        ).exclude(id__in=busy_ids).distinct()

        return Response(TeacherSerializer(available, many=True).data)


class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Audit loglari (faqat o'qish).

    Filtrlar (query params):
      ?action=create|update|delete|...
      ?model_name=Curriculum
      ?user_id=3
      ?date_from=YYYY-MM-DD
      ?date_to=YYYY-MM-DD
      ?search=... (object_repr / model_name ichidan qidiradi)
    """
    serializer_class = AuditLogSerializer
    permission_classes = [IsOrgAdmin]

    def get_queryset(self):
        qs = AuditLog.objects.filter(
            organization=self.request.user.organization
        ).select_related('user')

        params = self.request.query_params
        if action := params.get('action'):
            qs = qs.filter(action=action)
        if model_name := params.get('model_name'):
            qs = qs.filter(model_name=model_name)
        if user_id := params.get('user_id'):
            qs = qs.filter(user_id=user_id)
        if date_from := params.get('date_from'):
            qs = qs.filter(timestamp__date__gte=date_from)
        if date_to := params.get('date_to'):
            qs = qs.filter(timestamp__date__lte=date_to)
        if search := params.get('search'):
            qs = qs.filter(
                Q(object_repr__icontains=search) |
                Q(model_name__icontains=search)
            )
        return qs.order_by('-timestamp')

    @action(detail=False, methods=['get'], url_path='model-names')
    def model_names(self, request):
        """Filtr dropdown'i uchun — shu tashkilotda logga tushgan model nomlari."""
        names = (
            AuditLog.objects
            .filter(organization=request.user.organization)
            .values_list('model_name', flat=True)
            .distinct()
        )
        return Response(sorted(set(names)))


# ─────────────────────────────────────────────
#  TAQSIMOT VIEWSET
# ─────────────────────────────────────────────

class LoadSheetViewSet(viewsets.ModelViewSet):
    """
    Taqsimot varaqlari ro'yxati, ko'rish va o'chirish.
    Yuklash uchun: POST /load-sheets/upload/
    """
    http_method_names = ['get', 'post', 'delete', 'head', 'options']  # PUT/PATCH yo'q
    serializer_class   = LoadSheetSerializer
    permission_classes = [IsDeptManager]
    # MultiPartParser/FormParser — 'upload' (Excel fayl) uchun;
    # JSONParser — 'set-subject-teacher' (JSON body) uchun
    parser_classes     = [MultiPartParser, FormParser, JSONParser]

    def get_queryset(self):
        user = self.request.user
        qs   = LoadSheet.objects.filter(
            department__organization=user.organization
        ).select_related('department', 'uploaded_by').prefetch_related(
            'teacher_loads__distributions'
        )
        # Dept manager faqat o'z kafedrasi
        if user.role == 'dept_manager':
            dept = Department.objects.filter(
                manager=user, organization=user.organization
            ).first()
            if dept:
                qs = qs.filter(department=dept)
        return qs

    @action(detail=False, methods=['post'], url_path='upload',
            parser_classes=[MultiPartParser, FormParser])
    def upload(self, request):
        """
        POST /api/v1/load-sheets/upload/
        Body (form-data):
          - file: Excel fayl (.xlsx)
          - department_id: kafedra ID (ixtiyoriy, dept_manager uchun avtomatik)
          - year: yil (default: joriy yil)
        """
        file = request.FILES.get('file')
        if not file:
            return Response(
                {'error': 'Excel fayl yuklanmadi. "file" field talab qilinadi.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'Faqat .xlsx yoki .xls formatdagi fayllar qabul qilinadi.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        user = request.user

        # Kafedrani aniqlash
        dept_id = request.data.get('department_id')
        if dept_id:
            dept = Department.objects.filter(
                id=dept_id, organization=user.organization
            ).first()
            if not dept:
                return Response(
                    {'error': 'Kafedra topilmadi.'},
                    status=status.HTTP_404_NOT_FOUND
                )
        else:
            dept = Department.objects.filter(
                manager=user, organization=user.organization
            ).first()
            if not dept:
                return Response(
                    {'error': 'Siz hech qaysi kafedraning mudiri emassiz. department_id bering.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

        year = int(request.data.get('year', datetime.date.today().year))

        try:
            parsed = parse_load_sheet_excel(
                file=file,
                department=dept,
                year=year,
                organization=user.organization,
                uploaded_by=user,
            )
        except Exception as e:
            return Response(
                {'error': f'Excel faylni o\'qishda xato: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        results = parsed['sheets']
        skipped = parsed['skipped']

        if not results:
            return Response(
                {
                    'error': 'Excel faylda tan olinadigan oy sheeti topilmadi '
                             '(Yanvar, Fevral, ... Dekabr bo\'lishi kerak).',
                    'skipped': skipped,
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        return Response({
            'success':  True,
            'message':  f'{len(results)} ta oy muvaffaqiyatli yuklandi.',
            'sheets':   results,
            # O'tkazib yuborilgan sheet'lar — foydalanuvchi biror oy "yo'qolib qolgan"
            # deb o'ylamasligi uchun aniq ko'rsatiladi (haqiqiy bug tuzatildi:
            # avval bu haqda hech qanday ma'lumot qaytmasdi)
            'skipped':  skipped,
        }, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'], url_path='curriculum-preview',
            permission_classes=[IsAuthenticated])
    def curriculum_preview(self, request):
        """
        GET /api/v1/load-sheets/curriculum-preview/?month=M&year=YYYY

        O'sha oydagi guruhlarning o'quv rejasidagi fanlarni qaytaradi.
        - Guruhlar: GroupDayAssignment (kunlik kalendar) dan olinadi
        - dept_manager: faqat o'z kafedrasi ga biriktirilgan fanlar
        - Boshqa rollar: barcha fanlar
        """
        from academic.models import GroupDayAssignment, Curriculum

        month = request.query_params.get('month')
        year  = request.query_params.get('year')

        if not month or not year:
            return Response({'error': 'month va year talab qilinadi'}, status=400)
        try:
            month = int(month)
            year  = int(year)
        except ValueError:
            return Response({'error': 'month va year son bo\'lishi kerak'}, status=400)

        user = request.user
        org  = user.organization

        # Dept_manager bo'lsa — faqat o'z kafedrasi
        dept_filter = None
        if user.role == 'dept_manager':
            dept_filter = Department.objects.filter(
                manager=user, organization=org
            ).first()
            if not dept_filter:
                return Response({
                    'month': month, 'year': year,
                    'total_groups': 0, 'groups': [],
                    'warning': "Siz hech qaysi kafedraning mudiri emassiz.",
                })

        # O'sha oyda kunlik biriktiruvi bor guruhlar (takrorlanmasdan)
        group_ids = (
            GroupDayAssignment.objects
            .filter(
                group__organization=org,
                date__year=year,
                date__month=month,
            )
            .values_list('group_id', flat=True)
            .distinct()
        )

        if not group_ids:
            return Response({
                'month': month, 'year': year,
                'total_groups': 0, 'groups': [],
                'warning': (
                    f"{month}/{year} uchun guruh kunlik biriktiruvi topilmadi. "
                    "Avval 'Guruh biriktiruv' bo'limida kunlarga smena va bino biriktiring."
                ),
            })

        from academic.models import Group
        groups = (
            Group.objects
            .filter(id__in=group_ids, organization=org)
            .select_related('major')
            .order_by('name')
        )

        result = []
        for group in groups:
            if not group.major_id:
                continue

            is_online = group.delivery_mode == DeliveryMode.ONLINE

            # O'sha oydagi kunlik biriktiruv ma'lumotlari (birinchi yozuvdan smena/bino)
            first_da = (
                GroupDayAssignment.objects
                .filter(group=group, date__year=year, date__month=month)
                .select_related('shift', 'building')
                .order_by('date')
                .first()
            )
            shift_name = first_da.shift.name if first_da and first_da.shift else '—'
            # Onlayn guruhda bino umuman kerak emas — "Bino biriktirilmagan" (xato) bo'limi
            # o'rniga aniq "Onlayn (Zoom)" ko'rsatiladi (LoadSheetPage.jsx guruhlash uchun)
            building_name = (
                'Onlayn (Zoom)' if is_online
                else (first_da.building.name if first_da and first_da.building else '—')
            )

            # Faol o'quv rejasi — shu oy va shu guruhning turi (onlayn/oflayn) uchun
            # amal qiluvchisi (approved_date bo'yicha, academic/models.py::Curriculum.get_active_for_date)
            month_end_day = calendar.monthrange(year, month)[1]
            curriculum = Curriculum.get_active_for_date(
                group.major,
                target_date=datetime.date(year, month, month_end_day),
                delivery_mode=group.delivery_mode,
                queryset=Curriculum.objects.prefetch_related(
                    'blocks__subjects__subject',
                    'blocks__subjects__department',
                ),
            )

            if not curriculum:
                result.append({
                    'group_id':      group.id,
                    'group_name':    group.name,
                    'major_name':    group.major.name,
                    'shift_name':    shift_name,
                    'building_name': building_name,
                    'is_online':     is_online,
                    'curriculum':    None,
                    'subjects':      [],
                    'warning':       f"{group.major.name} uchun faol o'quv reja topilmadi.",
                })
                continue

            # Shu yo'nalishda qaysi o'qituvchi qaysi fanni o'ta oladi
            # (TeacherSubjectAssignment): subject_id -> [{id, name}]
            eligible_map = {}
            for tsa in (
                TeacherSubjectAssignment.objects
                .filter(major_id=group.major_id, teacher__organization=org)
                .select_related('teacher__user')
                .prefetch_related('subjects')
            ):
                tname = (tsa.teacher.user.get_full_name().strip()
                         or tsa.teacher.user.username)
                for subj in tsa.subjects.all():
                    eligible_map.setdefault(subj.id, []).append({
                        'id':   tsa.teacher_id,
                        'name': tname,
                    })

            # Shu GURUH uchun fanlarga biriktirilgan o'qituvchilar (GroupSubject) —
            # curriculum_subject_id -> GroupSubject. Har bir guruh o'z alohida
            # biriktiruvini oladi, xuddi shu o'quv rejani ishlatuvchi boshqa
            # guruhlarga ta'sir qilmaydi (haqiqiy, tuzatilgan bug —
            # load-sheet-teacher-assignment.md ga qarang).
            gs_by_cs = {
                gs.curriculum_subject_id: gs
                for gs in GroupSubject.objects.filter(
                    group=group, curriculum_subject__block__curriculum=curriculum
                ).select_related('teacher__user')
            }

            subjects = []
            has_any_subject = False
            has_any_department_assigned = False
            for block in curriculum.blocks.all():
                for cs in block.subjects.select_related('subject', 'department').all():
                    has_any_subject = True
                    if cs.department_id:
                        has_any_department_assigned = True
                    # dept_manager: faqat o'z kafedrasi fanlari
                    if dept_filter and cs.department_id != dept_filter.id:
                        continue

                    # O'qituvchi ma'lumoti — shu GURUH uchun (GroupSubject)
                    gs = gs_by_cs.get(cs.id)
                    teacher_id   = gs.teacher_id             if gs and gs.teacher else None
                    teacher_name = gs.teacher.user.get_full_name() if gs and gs.teacher else None
                    is_vacant    = bool(gs and gs.is_vacant)

                    subjects.append({
                        'curriculum_subject_id': cs.id,
                        'group_id':          group.id,
                        'module_number':     f"{block.order}.{cs.order}",
                        'block_name':        block.name or f'{block.order}-blok',
                        'subject_id':        cs.subject_id,
                        'subject_name':      cs.subject.name,
                        'subject_code':      cs.subject.code,
                        'department_id':     cs.department_id,
                        'department_name':   cs.department.name if cs.department else None,
                        'teacher_id':        teacher_id,
                        'teacher_name':      teacher_name,
                        'is_vacant':         is_vacant,
                        'eligible_teachers': eligible_map.get(cs.subject_id, []),
                        'lecture_hours':     cs.lecture_hours,
                        'practice_hours':    cs.practice_hours,
                        'field_hours':       cs.field_hours,
                        'independent_hours': cs.independent_hours,
                        'auditorium_hours':  cs.auditorium_hours,
                        'grand_total_hours': cs.grand_total_hours,
                        'week1_hours':       cs.week1_hours,
                        'week2_hours':       cs.week2_hours,
                        'week3_hours':       cs.week3_hours,
                        'week4_hours':       cs.week4_hours,
                    })

            # dept_manager va bu guruhda o'z kafedrasiga biriktirilgan fan yo'q bo'lsa:
            # - agar butun o'quv rejada UMUMAN hech qaysi fanga kafedra biriktirilmagan
            #   bo'lsa ("Fanlarga kafedra biriktirish" hali qilinmagan, masalan yangi
            #   Excel yuklangandan keyin) — guruh jimgina yashirilmaydi, aksincha aniq
            #   ogohlantirish bilan ko'rsatiladi. Aks holda (kafedralar biriktirilgan,
            #   lekin bu kafedraga tegishli fan yo'q) — guruh haqiqatan bu kafedraga
            #   tegishli emas, jimgina o'tkazib yuboriladi (avvalgi xatti-harakat).
            if dept_filter and not subjects:
                if has_any_subject and not has_any_department_assigned:
                    result.append({
                        'group_id':      group.id,
                        'group_name':    group.name,
                        'major_name':    group.major.name,
                        'shift_name':    shift_name,
                        'building_name': building_name,
                        'is_online':     is_online,
                        'curriculum':    curriculum.name,
                        'subjects':      [],
                        'warning': (
                            f"«{curriculum.name}» o'quv rejasidagi fanlarga hali kafedra "
                            "biriktirilmagan — 'O'quv reja' bo'limida \"Fanlarga kafedra "
                            "biriktirish\" orqali edu_admin/org_admin biriktirishi kerak."
                        ),
                    })
                continue

            result.append({
                'group_id':      group.id,
                'group_name':    group.name,
                'major_name':    group.major.name,
                'shift_name':    shift_name,
                'building_name': building_name,
                'is_online':     is_online,
                'curriculum':    curriculum.name,
                'subjects':      subjects,
                'warning':       None,
            })

        return Response({
            'month':        month,
            'year':         year,
            'total_groups': len(result),
            'groups':       result,
        })

    @action(detail=False, methods=['post'], url_path='set-subject-teacher',
            permission_classes=[IsAuthenticated])
    def set_subject_teacher(self, request):
        """
        POST /api/v1/load-sheets/set-subject-teacher/
        {
          "curriculum_subject_id": 12,
          "group_id": 7,
          "teacher_id": 5,         ← null bo'lsa biriktiruvni olib tashlaydi
          "is_vacant": true        ← ixtiyoriy, true bo'lsa "vakant" deb belgilaydi
        }

        Biriktiruv (GroupSubject) guruh + fan juftligiga tegishli — xuddi shu o'quv
        rejani ishlatuvchi BOSHQA guruhlarga ta'sir qilmaydi.

        `is_vacant=true` — hali o'qituvchi ishga olinmagan fanni belgilash uchun
        (`teacher_id` shu bilan birga bo'lmaydi). Keyinroq HAR QANDAY `teacher_id`
        bilan qayta chaqirilsa — `is_vacant` avtomatik `False`ga tushadi (vakansiya
        shu o'qituvchi bilan to'ldirildi deb hisoblanadi).
        """
        from academic.models import CurriculumSubject, Group

        cs_id      = request.data.get('curriculum_subject_id')
        group_id   = request.data.get('group_id')
        teacher_id = request.data.get('teacher_id')  # None = olib tashlash
        is_vacant  = bool(request.data.get('is_vacant'))

        if not cs_id:
            return Response({'error': 'curriculum_subject_id talab qilinadi'}, status=400)
        if not group_id:
            return Response({'error': 'group_id talab qilinadi'}, status=400)

        cs = CurriculumSubject.objects.filter(
            id=cs_id,
            block__curriculum__major__organization=request.user.organization,
        ).first()
        if not cs:
            return Response({'error': 'Fan topilmadi'}, status=404)

        group = Group.objects.filter(
            id=group_id, organization=request.user.organization,
        ).first()
        if not group:
            return Response({'error': 'Guruh topilmadi'}, status=404)

        if teacher_id:
            teacher = Teacher.objects.filter(
                id=teacher_id,
                organization=request.user.organization,
            ).first()
            if not teacher:
                return Response({'error': "O'qituvchi topilmadi"}, status=404)
            is_vacant = False   # o'qituvchi tanlansa vakansiya avtomatik to'ldiriladi
        else:
            teacher = None

        GroupSubject.objects.update_or_create(
            curriculum_subject=cs,
            group=group,
            defaults={
                'teacher':     teacher,
                'assigned_by': request.user,
                'is_vacant':   is_vacant,
            },
        )

        return Response({
            'curriculum_subject_id': cs_id,
            'group_id':     group.id,
            'teacher_id':   teacher.id            if teacher else None,
            'teacher_name': teacher.user.get_full_name() if teacher else None,
            'is_vacant':    is_vacant,
        })

    @action(detail=False, methods=['get'], url_path='template',
            permission_classes=[IsAuthenticated])
    def template(self, request):
        """
        GET /api/v1/load-sheets/template/
        Taqsimot Excel shablonini yuklab beradi.
        """
        wb = openpyxl.Workbook()

        MONTHS_UZ = [
            'Yanvar', 'Fevral', 'Mart', 'Aprel', 'May', 'Iyun',
            'Iyul', 'Avgust', 'Sentabr', 'Oktabr', 'Noyabr', 'Dekabr',
        ]

        header_fill  = PatternFill('solid', fgColor='1F4E79')
        header_font  = Font(bold=True, color='FFFFFF', size=10)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Side(style='thin', color='AAAAAA')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        example_groups = ['A-1', 'A-2', 'B-1', 'B-2']
        # Col layout: 0=T/r, 1=Modul, 2..5=Groups (4 ta misol), 6=Jami, 7=Hammasi, 8=Stavka, 9=FISh, 10=Lavozim
        # But we keep the real format: cols 2-22 for groups (21 max), 23=Jami, 24=Hammasi, 25=Stavka, 26=FISh, 27=Lavozim
        GROUP_START = 2
        GROUP_END   = 5   # 4 ta misol guruh (kengaytirish mumkin)
        COL_JAMI    = 23
        COL_HAMMASI = 24
        COL_STAVKA  = 25
        COL_FISH    = 26
        COL_LAVOZIM = 27
        TOTAL_COLS  = 28

        for month_name in MONTHS_UZ:
            ws = wb.create_sheet(title=month_name)

            # Qator 1 (index 0): Sarlavha
            ws.cell(row=1, column=1, value=f'{month_name} oyi — Taqsimot jadvali')
            ws.cell(row=1, column=1).font = Font(bold=True, size=12)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
            ws.cell(row=1, column=1).alignment = center_align

            # Qator 2 (index 1): Ustun nomlari
            headers = ['T/r', 'Modul nomi / Fan'] + \
                      [f'Guruh {i+1}' for i in range(21)] + \
                      ['Jami', 'Hammasi', 'Stavka', 'F.I.Sh', 'Lavozim']

            for col_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=2, column=col_idx, value=h)
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = center_align
                cell.border    = border

            # Misol qatorlar (index 2, 3)
            example_rows = [
                [1, '1.1. Jismoniy tayyorgarlik', 16, 8, 8, 8,
                 *([''] * 17), 40, 80, '1', 'Aliyev Vohid Rahimovich', 'Dotsent'],
                ['', '1.2. Nazariya', 8, 4, 4, 4,
                 *([''] * 17), 20, '', '', '', ''],
                [2, '2.1. Taktika', 12, 6, 6, 6,
                 *([''] * 17), 30, 60, '0.5', 'Karimov Jasur Baxtiyorovich', "O'qituvchi"],
            ]
            for r_idx, row_data in enumerate(example_rows, start=3):
                for c_idx, val in enumerate(row_data, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val if val != '' else None)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # Ustun kengliklari
            ws.column_dimensions['A'].width = 5    # T/r
            ws.column_dimensions['B'].width = 35   # Modul
            for col_letter in [ws.cell(row=1, column=c).column_letter for c in range(3, 24)]:
                ws.column_dimensions[col_letter].width = 8  # Guruhlar
            ws.column_dimensions[ws.cell(row=1, column=24).column_letter].width = 7   # Jami
            ws.column_dimensions[ws.cell(row=1, column=25).column_letter].width = 9   # Hammasi
            ws.column_dimensions[ws.cell(row=1, column=26).column_letter].width = 8   # Stavka
            ws.column_dimensions[ws.cell(row=1, column=27).column_letter].width = 25  # FISh
            ws.column_dimensions[ws.cell(row=1, column=28).column_letter].width = 15  # Lavozim

            ws.row_dimensions[2].height = 40

        # Birinchi bo'sh sheetni o'chirish
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="taqsimot_shablon.xlsx"'
        return response
