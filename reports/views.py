"""
Jadval hisobotlari:
  GET /api/v1/reports/schedule/<schedule_id>/excel/  → .xlsx
  GET /api/v1/reports/schedule/<schedule_id>/pdf/    → .pdf

Query params (ikkalasi uchun ham):
  ?group_id=X   — faqat bitta guruh (bo'lmasa barcha guruhlar)
"""

import io
import datetime
from collections import defaultdict

from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated

from openpyxl import Workbook
from openpyxl.styles import (Font, Alignment, PatternFill,
                              Border, Side, GradientFill)
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                 Paragraph, Spacer)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from scheduling.models import Schedule, ScheduleEntry


# ──────────────────────────────────────────────────────────────────────────────
#  YORDAMCHI FUNKSIYALAR
# ──────────────────────────────────────────────────────────────────────────────

WEEKDAYS_UZ = {
    1: 'dushanba', 2: 'seshanba', 3: 'chorshanba',
    4: 'payshanba', 5: 'juma', 6: 'shanba', 7: 'yakshanba',
}

MONTHS_UZ = {
    1: 'yanvar', 2: 'fevral', 3: 'mart', 4: 'aprel',
    5: 'may', 6: 'iyun', 7: 'iyul', 8: 'avgust',
    9: 'sentabr', 10: 'oktabr', 11: 'noyabr', 12: 'dekabr',
}

LESSON_TYPE_UZ = {'lecture': 'nazariy', 'practice': 'amaliy',
                  'field': "ko'chma", 'independent': 'mustaqil'}

ROMAN = {1: 'I', 2: 'II', 3: 'III', 4: 'IV',
         5: 'V', 6: 'VI', 7: 'VII', 8: 'VIII'}


def _get_entries(schedule: Schedule, group_id=None):
    """Jadval yozuvlarini guruh bo'yicha tartiblab qaytaradi."""
    qs = (
        ScheduleEntry.objects
        .filter(schedule=schedule)
        .select_related('teacher__user', 'group', 'subject', 'para', 'room', 'building')
        .order_by('group__name', 'date', 'para__order')
    )
    if group_id:
        qs = qs.filter(group_id=group_id)
    return qs


def _group_entries(entries):
    """
    {group: {date: [entry, ...]}} ko'rinishiga o'tkazadi.
    """
    result = defaultdict(lambda: defaultdict(list))
    for e in entries:
        result[e.group][e.date].append(e)
    return result


def _header_title(schedule: Schedule, group) -> str:
    org  = schedule.organization
    date_from = schedule.date_from
    date_to   = schedule.date_to
    return (
        f'"{org.name}" o\'quv mashg\'ulotlari jadvali\n'
        f'({date_from.strftime("%d.%m.%Y")}–{date_to.strftime("%d.%m.%Y")})'
    )


def _entry_rows(date_entries: dict) -> list[tuple]:
    """
    Har bir sana + entry dan jadval qatorlarini qaytaradi.
    (date_str, weekday, roman_para, time_range, module, lesson_type, teacher, room)
    """
    rows = []
    for date, entries in sorted(date_entries.items()):
        date_str = date.strftime('%d.%m.%Y')
        weekday  = WEEKDAYS_UZ.get(date.isoweekday(), '')
        for i, entry in enumerate(sorted(entries, key=lambda e: e.para.order)):
            roman    = ROMAN.get(entry.para.order, str(entry.para.order))
            time_str = (
                f"{entry.para.start_time.strftime('%H.%M')}-"
                f"{entry.para.end_time.strftime('%H.%M')}"
            )
            module   = entry.subject.name if entry.subject else '—'
            ltype    = LESSON_TYPE_UZ.get(entry.lesson_type, entry.lesson_type)
            teacher  = entry.teacher.user.get_full_name() if entry.teacher else '—'
            room     = entry.room.name if entry.room else (
                entry.building.name if entry.building else '—'
            )
            rows.append((
                date_str if i == 0 else '',   # birinchi parada sana ko'rinadi
                weekday  if i == 0 else '',
                roman,
                time_str,
                module,
                ltype,
                teacher,
                room,
            ))
    return rows


# ──────────────────────────────────────────────────────────────────────────────
#  EXCEL EXPORT
# ──────────────────────────────────────────────────────────────────────────────

def _thin_border():
    s = Side(style='thin', color='000000')
    return Border(left=s, right=s, top=s, bottom=s)


def _make_excel(schedule: Schedule, group_entries: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)          # bo'sh default sheetni o'chirish

    HEADER_FILL  = PatternFill('solid', fgColor='1F4E79')   # to'q ko'k
    HEADER_FONT  = Font(name='Times New Roman', bold=True, color='FFFFFF', size=10)
    TITLE_FONT   = Font(name='Times New Roman', bold=True, size=11)
    DATA_FONT    = Font(name='Times New Roman', size=10)
    DATE_FILL    = PatternFill('solid', fgColor='D6E4F0')    # och ko'k
    EVEN_FILL    = PatternFill('solid', fgColor='F2F2F2')    # kulrang
    CENTER       = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT         = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    BORDER       = _thin_border()

    COL_WIDTHS   = [12, 12, 7, 14, 42, 12, 28, 14]
    COL_HEADERS  = ['Sana', 'Hafta kuni', 'Juftlik',
                    'Vaqt', 'Modul nomi', 'Dars turi',
                    "O'qituvchi F.I.Sh.", 'Xona']

    for group, date_dict in sorted(group_entries.items(), key=lambda x: x[0].name):
        # Sheet nomi (Excel 31 belgidan uzun bo'lmaydi)
        sh_name = group.name[:31]
        ws = wb.create_sheet(title=sh_name)

        # ── 1. TASDIQLASH sarlavhasi ──────────────────────────────────────────
        ws.merge_cells('F1:H1')
        ws['F1'] = (
            '"TASDIQLAYMAN"\n'
            "O'quv ishlari bo'yicha prorektor\n"
            '____________\n'
            f'{schedule.date_from.year}-yil  {MONTHS_UZ[schedule.date_from.month]}'
        )
        ws['F1'].font      = Font(name='Times New Roman', italic=True, size=9)
        ws['F1'].alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
        ws.row_dimensions[1].height = 52

        # ── 2. SARLAVHA (kurs nomi + guruh + sana oralig'i) ──────────────────
        org       = schedule.organization
        date_from = schedule.date_from
        date_to   = schedule.date_to
        ws.merge_cells('A2:H2')
        ws['A2'] = (
            f'"{org.name}"\n'
            f'"{group.name}" guruh uchun o\'quv mashg\'ulotlari jadvali\n'
            f'({date_from.strftime("%d.%m.%Y")}–{date_to.strftime("%d.%m.%Y")})'
        )
        ws['A2'].font      = TITLE_FONT
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[2].height = 52

        # ── 3. USTUN SARLAVHALARI ─────────────────────────────────────────────
        for col_i, (header, width) in enumerate(zip(COL_HEADERS, COL_WIDTHS), start=1):
            cell = ws.cell(row=3, column=col_i, value=header)
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = CENTER
            cell.border    = BORDER
            ws.column_dimensions[get_column_letter(col_i)].width = width
        ws.row_dimensions[3].height = 30

        # ── 4. MA'LUMOT QATORLARI ─────────────────────────────────────────────
        rows = _entry_rows(date_dict)

        row_num = 4
        date_start_row = 4          # sana merge uchun
        prev_date = None

        for idx, (date_s, weekday, roman, time_s, module, ltype, teacher, room) in enumerate(rows):
            # Yangi sana boshlanganda — oldingi sanani merge qilish
            if date_s and prev_date is not None and row_num - 1 > date_start_row:
                ws.merge_cells(
                    start_row=date_start_row, start_column=1,
                    end_row=row_num - 1,      end_column=1
                )
                ws.merge_cells(
                    start_row=date_start_row, start_column=2,
                    end_row=row_num - 1,      end_column=2
                )

            if date_s:
                date_start_row = row_num
                prev_date = date_s

            row_fill = DATE_FILL if date_s else (EVEN_FILL if idx % 2 == 0 else None)
            values   = [date_s, weekday, roman, time_s, module, ltype, teacher, room]
            aligns   = [CENTER, CENTER, CENTER, CENTER, LEFT, CENTER, LEFT, CENTER]

            for col_i, (val, align) in enumerate(zip(values, aligns), start=1):
                cell = ws.cell(row=row_num, column=col_i, value=val)
                cell.font      = DATA_FONT
                cell.alignment = align
                cell.border    = BORDER
                if row_fill:
                    cell.fill = row_fill

            ws.row_dimensions[row_num].height = 22
            row_num += 1

        # Oxirgi sana merge
        if row_num - 1 >= date_start_row + 1:
            ws.merge_cells(
                start_row=date_start_row, start_column=1,
                end_row=row_num - 1,      end_column=1
            )
            ws.merge_cells(
                start_row=date_start_row, start_column=2,
                end_row=row_num - 1,      end_column=2
            )

        # ── 5. IMZO QATORI ────────────────────────────────────────────────────
        ws.merge_cells(f'A{row_num + 1}:H{row_num + 1}')
        ws.cell(row=row_num + 1, column=1,
                value="Kafedralar mudiri: _______________").font = DATA_FONT

        # Sahifani gorizontal chop etishga sozlash
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.paperSize   = ws.PAPERSIZE_A4
        ws.print_title_rows       = '1:3'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────────────────────
#  PDF EXPORT  (reportlab)
# ──────────────────────────────────────────────────────────────────────────────

def _make_pdf(schedule: Schedule, group_entries: dict) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    styles  = getSampleStyleSheet()
    story   = []

    title_style = ParagraphStyle(
        'Title', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=11,
        alignment=TA_CENTER, spaceAfter=6,
    )
    small_style = ParagraphStyle(
        'Small', parent=styles['Normal'],
        fontName='Helvetica', fontSize=8,
        alignment=TA_LEFT,
    )
    cell_style = ParagraphStyle(
        'Cell', parent=styles['Normal'],
        fontName='Helvetica', fontSize=8,
        alignment=TA_LEFT,
    )
    center_style = ParagraphStyle(
        'Center', parent=styles['Normal'],
        fontName='Helvetica', fontSize=8,
        alignment=TA_CENTER,
    )

    org       = schedule.organization
    date_from = schedule.date_from
    date_to   = schedule.date_to

    # Ustun kengliklari (cm → points: 1cm ≈ 28.35pt)
    # A4 landscape: ~27cm qo'llanma kengligi
    COL_W = [2.2*cm, 2.2*cm, 1.3*cm, 2.4*cm, 8*cm, 2.2*cm, 5*cm, 2.4*cm]
    HEADERS = ['Sana', 'Hafta kuni', 'Juft.', 'Vaqt',
               'Modul nomi', 'Dars turi', "O'qituvchi", 'Xona']

    HDR_BG    = colors.HexColor('#1F4E79')
    DATE_BG   = colors.HexColor('#D6E4F0')
    EVEN_BG   = colors.HexColor('#F2F2F2')
    WHITE     = colors.white
    BLACK     = colors.black

    for group, date_dict in sorted(group_entries.items(), key=lambda x: x[0].name):
        # Guruh sarlavhasi
        story.append(Paragraph(
            f'"{org.name}" — <b>{group.name}</b> guruh jadvali',
            title_style,
        ))
        story.append(Paragraph(
            f'{date_from.strftime("%d.%m.%Y")} – {date_to.strftime("%d.%m.%Y")}',
            ParagraphStyle('Sub', parent=styles['Normal'],
                           fontName='Helvetica', fontSize=9, alignment=TA_CENTER),
        ))
        story.append(Spacer(1, 0.3 * cm))

        rows_data = _entry_rows(date_dict)

        # Jadval ma'lumotlari
        table_data = [HEADERS]  # birinchi qator — sarlavha

        # Har qator uchun ranglar
        row_colors = []      # (row_idx, color)
        merge_spans = []     # (row_i, col, rowspan) — reportlab SPAN

        cur_table_row = 1    # 0-sarlavha
        date_span_start = 1
        prev_date = None

        for idx, (date_s, weekday, roman, time_s, module, ltype, teacher, room) in enumerate(rows_data):
            if date_s and prev_date is not None:
                # Oldingi sanani span qilish
                span = cur_table_row - date_span_start
                if span > 1:
                    merge_spans.append(('SPAN', (0, date_span_start), (0, cur_table_row - 1)))
                    merge_spans.append(('SPAN', (1, date_span_start), (1, cur_table_row - 1)))
                row_colors.append((date_span_start, DATE_BG))
                date_span_start = cur_table_row
                prev_date = date_s
            elif date_s:
                prev_date = date_s
                date_span_start = cur_table_row

            bg = EVEN_BG if idx % 2 == 0 else WHITE

            table_data.append([
                Paragraph(date_s, center_style),
                Paragraph(weekday, center_style),
                Paragraph(roman, center_style),
                Paragraph(time_s, center_style),
                Paragraph(module, cell_style),
                Paragraph(ltype, center_style),
                Paragraph(teacher, cell_style),
                Paragraph(room, center_style),
            ])
            cur_table_row += 1

        # Oxirgi sana span
        if rows_data:
            span = cur_table_row - date_span_start
            if span > 1:
                merge_spans.append(('SPAN', (0, date_span_start), (0, cur_table_row - 1)))
                merge_spans.append(('SPAN', (1, date_span_start), (1, cur_table_row - 1)))
            row_colors.append((date_span_start, DATE_BG))

        t = Table(table_data, colWidths=COL_W, repeatRows=1)

        ts = TableStyle([
            # Sarlavha
            ('BACKGROUND',  (0, 0), (-1, 0),  HDR_BG),
            ('TEXTCOLOR',   (0, 0), (-1, 0),  WHITE),
            ('FONTNAME',    (0, 0), (-1, 0),  'Helvetica-Bold'),
            ('FONTSIZE',    (0, 0), (-1, 0),  8),
            ('ALIGN',       (0, 0), (-1, 0),  'CENTER'),
            ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
            # Grid
            ('GRID',        (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE',    (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [WHITE, EVEN_BG]),
        ])

        # Sana rangi
        for row_i, color in row_colors:
            ts.add('BACKGROUND', (0, row_i), (1, row_i), DATE_BG)

        # Merge
        for span_cmd in merge_spans:
            ts.add(*span_cmd)

        t.setStyle(ts)
        story.append(t)
        story.append(Spacer(1, 0.8 * cm))

        # Imzo
        story.append(Paragraph("Kafedralar mudiri: _______________", small_style))
        story.append(Spacer(1, 1 * cm))

    doc.build(story)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────────────────────
#  API VIEWS
# ──────────────────────────────────────────────────────────────────────────────

class ScheduleExcelExportView(APIView):
    """
    GET /api/v1/reports/schedule/<schedule_id>/excel/
    Query params: ?group_id=X (ixtiyoriy)

    Faqat o'z tashkiloti jadvalini yuklab olish mumkin.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, schedule_id):
        schedule = (
            Schedule.objects
            .filter(id=schedule_id,
                    organization=request.user.organization)
            .select_related('organization')
            .first()
        )
        if not schedule:
            from rest_framework.response import Response
            from rest_framework import status as drf_status
            return Response({'error': 'Jadval topilmadi.'}, status=drf_status.HTTP_404_NOT_FOUND)

        group_id = request.query_params.get('group_id')
        entries  = _get_entries(schedule, group_id)
        grouped  = _group_entries(entries)

        if not grouped:
            from rest_framework.response import Response
            from rest_framework import status as drf_status
            return Response({'error': 'Jadval yozuvlari yo\'q.'}, status=drf_status.HTTP_404_NOT_FOUND)

        xlsx_bytes = _make_excel(schedule, grouped)

        month_name = MONTHS_UZ.get(schedule.month, str(schedule.month))
        filename   = f"jadval_{month_name}_{schedule.year}.xlsx"

        response = HttpResponse(
            xlsx_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class SchedulePDFExportView(APIView):
    """
    GET /api/v1/reports/schedule/<schedule_id>/pdf/
    Query params: ?group_id=X (ixtiyoriy)
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, schedule_id):
        schedule = (
            Schedule.objects
            .filter(id=schedule_id,
                    organization=request.user.organization)
            .select_related('organization')
            .first()
        )
        if not schedule:
            from rest_framework.response import Response
            from rest_framework import status as drf_status
            return Response({'error': 'Jadval topilmadi.'}, status=drf_status.HTTP_404_NOT_FOUND)

        group_id = request.query_params.get('group_id')
        entries  = _get_entries(schedule, group_id)
        grouped  = _group_entries(entries)

        if not grouped:
            from rest_framework.response import Response
            from rest_framework import status as drf_status
            return Response({'error': 'Jadval yozuvlari yo\'q.'}, status=drf_status.HTTP_404_NOT_FOUND)

        pdf_bytes = _make_pdf(schedule, grouped)

        month_name = MONTHS_UZ.get(schedule.month, str(schedule.month))
        filename   = f"jadval_{month_name}_{schedule.year}.pdf"

        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
